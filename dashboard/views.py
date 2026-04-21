import json
from datetime import datetime, time
import io
import threading
import csv
import os
import tempfile
import shutil
import mimetypes
from urllib.parse import unquote
import time
import smtplib
import ssl
from django.conf import settings
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.mail import EmailMultiAlternatives
from django.core.management import call_command
from django.core.paginator import Paginator
from django.core.files.storage import default_storage
from django import forms
from django.core.exceptions import ValidationError
from django.db import IntegrityError
from django.db import transaction
from django.db.models import Count, Max, Q, Avg, F, OuterRef, Subquery
from django.db.models.functions import Trim
from django.http import HttpResponseForbidden, HttpResponse, Http404, FileResponse
from django.http import JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone
from django.template.loader import render_to_string
from django.http import JsonResponse

from fcm_django.models import FCMDevice

from attendance.models import Attendance
from groups.models import Group
from students.models import Student, StudentRegistration, Badge, StudentBadge, StudentEventStats
from users.models import User
from workshops.models import Workshop, WorkshopSession, WorkshopFeedback
from api.models import MobileDevice
from public_screen.models import PublicForm, PublicFormSubmission, PublicFormAnswer, PublicFormField

from .models import (
    AdminLog, Notification, Event, VIPInvite, VolunteerNote, 
    StudentViolation, FailedEmail, SOSRequest, BroadcastMessage, 
    StudentSupportRequest, AppVersion
)
from techday.utils import send_email_async, get_styled_email_html


@login_required
def admin_broadcast_send(request):
    """
    Send a global broadcast message (AJAX)
    """
    if not require_admin(request.user):
        return JsonResponse({'success': False}, status=403)
    
    if request.method == 'POST':
        message = request.POST.get('message', '').strip()
        target = request.POST.get('target', 'all')
        duration = int(request.POST.get('duration', 5)) # minutes
        
        if not message:
            return JsonResponse({'success': False, 'error': 'يجب كتابة نص الإذاعة'}, status=400)
            
        expires_at = timezone.now() + timezone.timedelta(minutes=duration)
        
        # Deactivate previous active broadcasts for the same target to avoid clutter
        BroadcastMessage.objects.filter(target=target, is_active=True).update(is_active=False)
        
        BroadcastMessage.objects.create(
            event=Event.get_current(),
            author=request.user,
            message=message,
            target=target,
            expires_at=expires_at
        )
        return JsonResponse({'success': True})
    return JsonResponse({'success': False}, status=400)


@login_required
def user_broadcast_poll(request):
    """
    Poll for active broadcasts targeted at the user (AJAX)
    """
    now = timezone.now()
    
    # Determine user category
    user_target = 'all'
    if hasattr(request.user, 'role'):
        if request.user.role == 'supervisor':
            user_target = 'supervisors'
        elif request.user.role == 'volunteer':
            user_target = 'volunteers'
    
    # Check if student (no role but might be authenticated)
    if not hasattr(request.user, 'role') or request.user.role == 'student':
        if Student.objects.filter(user=request.user).exists():
            user_target = 'students'

    # Get active broadcasts for 'all' OR the specific user category
    broadcasts = BroadcastMessage.objects.filter(
        Q(target='all') | Q(target=user_target),
        is_active=True,
        expires_at__gt=now
    ).order_by('-created_at')
    
    data = []
    for b in broadcasts:
        data.append({
            'id': b.id,
            'message': b.message,
            'target': b.get_target_display(),
            'type': b.target
        })
    
    return JsonResponse({'success': True, 'broadcasts': data})


def require_admin(user):
    return user.is_authenticated and hasattr(user, 'is_admin') and user.is_admin()


@login_required
def admin_sos_poll(request):
    """
    Poll for new SOS requests (AJAX)
    """
    if not require_admin(request.user):
        return JsonResponse({'success': False}, status=403)
    
    # Get unread SOS requests
    new_requests = SOSRequest.objects.filter(is_seen=False, status=SOSRequest.Status.PENDING).select_related('supervisor', 'workshop', 'student')
    data = []
    for req in new_requests:
        location_text = req.workshop.room if req.workshop else (req.location_manual or 'غير محدد')
        student_text = f"الطالب: {req.student.name} ({req.student.student_id})" if req.student else ""
        
        data.append({
            'id': req.id,
            'supervisor': req.supervisor.get_full_name() or req.supervisor.username,
            'workshop': req.workshop.title if req.workshop else '',
            'room': location_text,
            'student': student_text,
            'message': req.message,
            'time': req.created_at.strftime('%H:%M') if req.created_at else timezone.now().strftime('%H:%M'),
        })
    
    return JsonResponse({'success': True, 'requests': data})


@login_required
def admin_sos_mark_seen(request):
    """
    Mark an SOS request as seen and optionally reply (AJAX)
    """
    if not require_admin(request.user):
        return JsonResponse({'success': False}, status=403)
    
    if request.method == 'POST':
        sos_id = request.POST.get('id')
        reply = request.POST.get('reply', '').strip()
        
        update_data = {'is_seen': True}
        if reply:
            update_data['admin_reply'] = reply
            update_data['reply_at'] = timezone.now()
            update_data['status'] = SOSRequest.Status.SOLVED
            
        SOSRequest.objects.filter(id=sos_id).update(**update_data)
        return JsonResponse({'success': True})
    return JsonResponse({'success': False}, status=400)


@login_required
def user_sos_poll_replies(request):
    """
    Poll for replies to user's own SOS requests (AJAX)
    """
    # Get unread replies for the logged in user
    new_replies = SOSRequest.objects.filter(
        supervisor=request.user,
        admin_reply__isnull=False,
        is_reply_seen=False
    ).exclude(admin_reply='').order_by('-reply_at')
    
    data = []
    for req in new_replies:
        data.append({
            'id': req.id,
            'reply': req.admin_reply,
            'time': req.reply_at.strftime('%H:%M') if req.reply_at else '',
            'original_message': req.message[:50] + '...' if len(req.message) > 50 else req.message
        })
    
    return JsonResponse({'success': True, 'replies': data})


@login_required
def user_sos_mark_reply_seen(request):
    """
    Mark an SOS reply as seen by the user (AJAX)
    """
    if request.method == 'POST':
        sos_id = request.POST.get('id')
        SOSRequest.objects.filter(id=sos_id, supervisor=request.user).update(is_reply_seen=True)
        return JsonResponse({'success': True})
    return JsonResponse({'success': False}, status=400)


@login_required
def supervisor_sos_send(request):
    """
    Send an SOS request (AJAX) - Automatically detects current workshop for supervisors
    And handles manual locations/students for volunteers
    """
    is_admin = require_admin(request.user)
    is_volunteer = hasattr(request.user, 'role') and request.user.role == 'volunteer'
    is_supervisor = hasattr(request.user, 'role') and request.user.role == 'supervisor'

    if not (is_supervisor or is_volunteer or is_admin):
        return JsonResponse({'success': False}, status=403)
    
    if request.method == 'POST':
        message = request.POST.get('message', '').strip()
        student_id = request.POST.get('student_id')
        location_manual = request.POST.get('location_manual', '').strip()
        now = timezone.localtime()
        
        if not message:
            return JsonResponse({'success': False, 'error': 'يجب كتابة نص الاستغاثة'}, status=400)
            
        workshop = None
        student = None

        # 1. التعرف على الطالب إذا تم إرساله
        if student_id:
            student = Student.objects.filter(Q(student_id=student_id) | Q(id=student_id)).first()

        # 2. التعرف على الورشة (للمشرفين فقط تلقائياً)
        if is_supervisor:
            current_session = WorkshopSession.objects.filter(
                workshop__supervisor=request.user,
                start_time__lte=now.time(),
                end_time__gte=now.time()
            ).select_related('workshop').first()
            
            if current_session:
                workshop = current_session.workshop
            else:
                workshop = Workshop.objects.filter(supervisor=request.user).first()
            
        SOSRequest.objects.create(
            supervisor=request.user,
            workshop=workshop,
            student=student,
            location_manual=location_manual,
            message=message,
            created_at=timezone.now()
        )
        return JsonResponse({'success': True})
    return JsonResponse({'success': False}, status=400)


@login_required
def admin_archived_events_list(request):
    if not require_admin(request.user):
        return HttpResponseForbidden()
    
    active_event = Event.get_current()
    archived_events = Event.objects.filter(is_archived=True).order_by('-created_at')
    
    return render(request, 'dashboard/admin_archived_events_list.html', {
        'active_event': active_event,
        'archived_events': archived_events
    })


@login_required
def admin_archived_event_detail(request, slug):
    # This will show the frozen data for a specific event
    event = get_object_or_404(Event, slug=slug)
    
    # Check if the requester is a student or staff
    student = getattr(request.user, 'student_profile', None)
    is_staff = request.user.is_staff or (hasattr(request.user, 'is_admin') and request.user.is_admin())
    
    # Fetch data linked to THIS event
    context = {
        'event': event,
        'is_archive_view': True,
        'is_staff': is_staff,
        'total_students': StudentEventStats.objects.filter(event=event).count(),
        'total_present': StudentEventStats.objects.filter(event=event, checked_in=True).count(),
        'workshops': Workshop.objects.filter(event=event),
        'groups': Group.objects.filter(event=event).order_by('-points'),
        'top_stats': StudentEventStats.objects.filter(event=event, points__gt=0).select_related('student', 'student__group').order_by('-points')[:10],
    }

    if is_staff:
        context['recent_logs'] = AdminLog.objects.filter(event=event)[:20]
    elif student:
        # Student specific data for this archived event
        stats = StudentEventStats.objects.filter(student=student, event=event).first()
        if stats:
            context['student_stats'] = stats
            context['student_rank'] = StudentEventStats.objects.filter(event=event, points__gt=stats.points).count() + 1
        
        # Student's personal logs (awards/violations) for this event
        personal_logs = AdminLog.objects.filter(
            event=event, 
            action__icontains=student.name
        ).order_by('-created_at')
        context['personal_logs'] = personal_logs

    template_name = 'dashboard/admin_archived_event_detail.html' if is_staff else 'dashboard/student_archived_event_detail.html'
    return render(request, template_name, context)


@login_required
def admin_dashboard(request):
    if not require_admin(request.user):
        return HttpResponseForbidden()
    
    event = Event.get_current()
        
    now = timezone.localtime()
    today = now.date()
    
    # Statistics should be for the CURRENT event
    # Total Students = Approved registrations for this event
    total_students = StudentRegistration.objects.filter(
        event=event,
        status=StudentRegistration.Status.APPROVED,
        removed_at__isnull=True,
    ).count()
    total_present = Student.objects.filter(checked_in=True).count()
    
    # General System Stats (Accounts)
    students_with_accounts = Student.objects.filter(user__isnull=False).count()
    total_logins_ever = Student.objects.filter(user__last_login__isnull=False).count()
    total_logins_today = Student.objects.filter(user__last_login__date=today).count()
    
    workshops = Workshop.objects.filter(event=event)
    groups = Group.objects.filter(event=event)
    active_workshops = workshops.filter(status='active').count()
    latest_notification = Notification.objects.filter(event=event, is_active=True).first()
    recent_logs = AdminLog.objects.filter(event=event)[:5]
    
    pending_emails_count = FailedEmail.objects.count()
    today_notes_count = VolunteerNote.objects.filter(created_at__date=today).count()
    
    event_status = 'لم تبدأ بعد'
    if event.is_finished:
        event_status = 'متوقفة'
    elif event.start_datetime:
        if now < event.start_datetime:
            event_status = 'لم تبدأ بعد'
        elif now >= event.start_datetime:
            event_status = 'جارية'
            
    context = {
        'now': now,
        'total_students': total_students,
        'total_present': total_present,
        'students_with_accounts': students_with_accounts,
        'total_logins_ever': total_logins_ever,
        'total_logins_today': total_logins_today,
        'workshops': workshops,
        'groups': groups,
        'active_workshops': active_workshops,
        'latest_notification': latest_notification,
        'recent_logs': recent_logs,
        'event_status': event_status,
        'event': event,
        'pending_emails_count': pending_emails_count,
        'today_notes_count': today_notes_count,
    }
    return render(request, 'dashboard/admin_dashboard.html', context)


@login_required
def admin_event_settings(request):
    if not require_admin(request.user):
        return HttpResponseForbidden()
    event = Event.get_current()
    if not event:
        event = Event.objects.create(name="Tech Day", location_name="Main", year=2026, is_active=True)
    if request.method == 'POST':
        action = request.POST.get('action') or 'save'
        if action == 'config':
            is_registration_closed = request.POST.get('is_registration_closed') == 'on'
            allow_existing = request.POST.get('allow_existing_students_registration') == 'on'
            allow_group_points = request.POST.get('allow_group_points') == 'on'
            is_maintenance_mode = request.POST.get('is_maintenance_mode') == 'on'
            is_admin_locked = request.POST.get('is_education_admin_locked') == 'on'
            locked_admin_list = request.POST.getlist('locked_education_admin')
            locked_admin = ','.join(locked_admin_list).strip()
            
            if is_admin_locked and not locked_admin:
                locked_admin = 'العبور'

            event.is_registration_closed = is_registration_closed
            event.allow_existing_students_registration = allow_existing
            event.allow_group_points = allow_group_points
            event.is_maintenance_mode = is_maintenance_mode
            event.is_education_admin_locked = is_admin_locked
            event.locked_education_admin = locked_admin
            event.save(update_fields=[
                'is_registration_closed',
                'allow_existing_students_registration',
                'allow_group_points',
                'is_maintenance_mode',
                'is_education_admin_locked',
                'locked_education_admin',
            ])
            AdminLog.objects.create(action='تم تحديث لوحة التحكم بالقيود (Config Panel)')
            messages.success(request, 'تم حفظ إعدادات القيود بنجاح.')
            return redirect('dashboard:admin_event_settings')
        if action == 'end':
            Workshop.objects.exclude(status='finished').update(status='finished')
            event.is_finished = True
            event.save(update_fields=['is_finished'])
            AdminLog.objects.create(action='تم إنهاء الفعالية وتم تعليم جميع الورش كمنتهية')
            messages.success(request, 'تم إنهاء الفعالية وتم تعليم جميع الورش كمنتهية.')
            return redirect('dashboard:admin_event_settings')
        if action == 'resume':
            event.is_finished = False
            event.save(update_fields=['is_finished'])
            AdminLog.objects.create(action='تم إعادة تشغيل الفعالية مرة أخرى بعد إنهائها')
            messages.success(request, 'تم إعادة تشغيل الفعالية مرة أخرى.')
            return redirect('dashboard:admin_event_settings')
        start_date = request.POST.get('start_date') or ''
        start_time = request.POST.get('start_time') or ''
        location_name = (request.POST.get('location_name') or '').strip()
        location_link = (request.POST.get('location_link') or '').strip()
        arrival_time_text = (request.POST.get('arrival_time_text') or '').strip()
        whatsapp_group_link = (request.POST.get('whatsapp_group_link') or '').strip()
        event_instructions = (request.POST.get('event_instructions') or '').strip()
        max_students_raw = (request.POST.get('max_students') or '').strip()
        max_students = int(max_students_raw) if max_students_raw.isdigit() else None

        if start_date and start_time:
            try:
                naive = datetime.strptime(f'{start_date} {start_time}', '%Y-%m-%d %H:%M')
                event.start_datetime = timezone.make_aware(naive, timezone.get_current_timezone())
            except ValueError:
                event.start_datetime = None
                messages.error(request, 'صيغة التاريخ أو الوقت غير صحيحة.')
        else:
            event.start_datetime = None
        event.location_name = location_name
        event.location_link = location_link
        event.arrival_time_text = arrival_time_text
        event.whatsapp_group_link = whatsapp_group_link
        event.event_instructions = event_instructions
        event.max_students = max_students
        event.save()
        if not messages.get_messages(request):
            messages.success(request, 'تم حفظ إعدادات الفعالية.')
        return redirect('dashboard:admin_event_settings')
    total_registered = StudentRegistration.objects.filter(
        event=event,
        status=StudentRegistration.Status.APPROVED,
        removed_at__isnull=True,
    ).count()
    is_full = event.max_students is not None and total_registered >= event.max_students
    
    # تحويل السلسلة النصية إلى قائمة للإدارات المقفولة
    locked_admins_list = [a.strip() for a in (event.locked_education_admin or '').split(',') if a.strip()]
    
    education_admin_choices = [
        'بنها', 'طوخ', 'كفر شكر', 'شبين القناطر', 'الخانكة', 'قها',
        'قليوب', 'القناطر الخيرية', 'غرب شبرا الخيمة', 'شرق شبرا الخيمة',
        'الخصوص', 'العبور',
    ]
    return render(request, 'dashboard/admin_event_settings.html', {
        'event': event,
        'total_registered': total_registered,
        'is_full': is_full,
        'education_admin_choices': education_admin_choices,
        'locked_admins_list': locked_admins_list,
    })


@login_required
def admin_archive_event(request):
    if not require_admin(request.user):
        return HttpResponseForbidden()
    
    current_event = Event.get_current()
    if not current_event:
        messages.error(request, 'لا توجد فعالية نشطة حالياً لأرشفتها.')
        return redirect('dashboard:admin_event_settings')
    
    if request.method == 'POST':
        with transaction.atomic():
            # 1. Capture current student stats into StudentEventStats
            students = Student.objects.all()
            for student in students:
                StudentEventStats.objects.update_or_create(
                    student=student,
                    event=current_event,
                    defaults={
                        'points': student.points,
                        'checked_in': student.checked_in,
                        'checked_in_at': student.checked_in_at,
                    },
                )
                # Reset student for next event
                student.points = 0
                student.checked_in = False
                student.checked_in_at = None
                student.save(update_fields=['points', 'checked_in', 'checked_in_at'])
            
            # 2. Archive the event
            current_event.is_active = False
            current_event.is_archived = True
            current_event.is_finished = True
            current_event.save()
            
            # 3. Create a new empty event
            # Logic for execution number: same location, same year -> increment
            next_exec = Event.objects.filter(
                location_name=current_event.location_name,
                year=current_event.year
            ).count() + 1
            
            Event.objects.create(
                name=current_event.name,
                location_name=current_event.location_name,
                location_link=current_event.location_link,
                year=current_event.year,
                execution_number=next_exec,
                arrival_time_text=current_event.arrival_time_text,
                whatsapp_group_link=current_event.whatsapp_group_link,
                event_instructions=current_event.event_instructions,
                max_students=current_event.max_students,
                is_active=True
            )
            
            AdminLog.objects.create(
                action=f'تم أرشفة الفعالية {current_event} وبدء فعالية جديدة رقم {next_exec}'
            )
            
            messages.success(request, f'تم أرشفة الفعالية بنجاح برابط: {current_event.slug}. النظام الآن جاهز للفعالية الجديدة.')
            return redirect('dashboard:admin_archived_events_list')
            
    return render(request, 'dashboard/admin_archive_confirm.html', {'event': current_event})


@login_required
def admin_toggle_group_points(request):
    if not require_admin(request.user):
        return JsonResponse({'success': False, 'error': 'غير مصرح'})
    
    event = Event.get_current()
    event.allow_group_points = not event.allow_group_points
    event.save()
    
    status = "مفعلة" if event.allow_group_points else "مغلقة"
    AdminLog.objects.create(
        action=f'تم تغيير حالة إضافة النقاط للمجموعات إلى: {status}'
    )
    
    return JsonResponse({
        'success': True, 
        'is_allowed': event.allow_group_points,
        'message': f'تم {status} إضافة النقاط بنجاح.'
    })


def check_app_version(request):
    """
    API to check the app version and return update info (Public API)
    """
    platform = request.GET.get('platform', 'android').lower()
    try:
        current_build = int(request.GET.get('build', 0))
    except (ValueError, TypeError):
        current_build = 0
    
    latest_version = AppVersion.objects.filter(platform=platform).order_by('-build_number').first()
    
    if not latest_version:
        return JsonResponse({
            'update_required': False,
            'update_available': False,
        })
        
    update_available = latest_version.build_number > current_build
    update_required = (
        latest_version.min_build_number > current_build
        or (latest_version.is_force_update and update_available)
    )

    download_url = latest_version.download_url
    if not download_url and latest_version.apk_file:
        download_url = request.build_absolute_uri(latest_version.apk_file.url)
    
    return JsonResponse({
        'update_available': update_available,
        'update_required': update_required,
        'is_force_update': latest_version.is_force_update and update_available,
        'latest_version': latest_version.version_code,
        'latest_build': latest_version.build_number,
        'download_url': download_url,
        'release_notes': latest_version.release_notes,
    })


class AppVersionForm(forms.ModelForm):
    class Meta:
        model = AppVersion
        fields = [
            'platform',
            'version_code',
            'build_number',
            'min_build_number',
            'download_url',
            'apk_file',
            'release_notes',
            'is_force_update',
        ]
        widgets = {
            'platform': forms.Select(),
            'version_code': forms.TextInput(),
            'build_number': forms.NumberInput(),
            'min_build_number': forms.NumberInput(),
            'download_url': forms.URLInput(),
            'apk_file': forms.ClearableFileInput(),
            'release_notes': forms.Textarea(attrs={'rows': 5}),
            'is_force_update': forms.CheckboxInput(),
        }

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        input_class = (
            'w-full px-4 py-3 rounded-2xl bg-slate-950 border border-slate-800 text-slate-200 '
            'focus:outline-none focus:ring-2 focus:ring-cyan-500/50 focus:border-cyan-500 transition-all font-bold text-sm'
        )
        for name, field in self.fields.items():
            if isinstance(field.widget, forms.CheckboxInput):
                field.widget.attrs.update({
                    'class': 'w-5 h-5 rounded border-slate-700 bg-slate-950 text-cyan-500 focus:ring-cyan-500/50',
                })
            elif isinstance(field.widget, (forms.Textarea,)):
                field.widget.attrs.update({
                    'class': input_class,
                })
            else:
                field.widget.attrs.update({
                    'class': input_class,
                })

    def clean(self):
        cleaned = super().clean()
        platform = cleaned.get('platform')
        download_url = (cleaned.get('download_url') or '').strip()
        apk_file = cleaned.get('apk_file')
        if platform == 'android' and not download_url and not apk_file:
            raise ValidationError('Android لازم يكون له إما رابط تحميل أو ملف APK.')
        return cleaned


@login_required
def admin_app_versions(request):
    if not require_admin(request.user):
        return HttpResponseForbidden()

    platform = (request.GET.get('platform') or '').strip().lower()
    qs = AppVersion.objects.all().order_by('-created_at', '-build_number')
    if platform in ('android', 'ios'):
        qs = qs.filter(platform=platform)

    return render(request, 'dashboard/admin_app_versions.html', {
        'versions': list(qs),
        'platform': platform,
    })


@login_required
def admin_app_version_create(request):
    if not require_admin(request.user):
        return HttpResponseForbidden()

    if request.method == 'POST':
        form = AppVersionForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                version = form.save()
                AdminLog.objects.create(action=f'تم إضافة إصدار تطبيق: {version}')
                messages.success(request, 'تم إضافة الإصدار بنجاح.')
                return redirect('dashboard:admin_app_versions')
            except IntegrityError:
                form.add_error(None, 'نفس (Platform + Build Number) موجود بالفعل.')
    else:
        form = AppVersionForm()

    return render(request, 'dashboard/admin_app_version_form.html', {
        'form': form,
        'version_obj': None,
    })


@login_required
def admin_app_version_update(request, pk):
    if not require_admin(request.user):
        return HttpResponseForbidden()

    version_obj = get_object_or_404(AppVersion, pk=pk)
    if request.method == 'POST':
        form = AppVersionForm(request.POST, request.FILES, instance=version_obj)
        if form.is_valid():
            try:
                version = form.save()
                AdminLog.objects.create(action=f'تم تعديل إصدار تطبيق: {version}')
                messages.success(request, 'تم حفظ التعديلات بنجاح.')
                return redirect('dashboard:admin_app_versions')
            except IntegrityError:
                form.add_error(None, 'نفس (Platform + Build Number) موجود بالفعل.')
    else:
        form = AppVersionForm(instance=version_obj)

    return render(request, 'dashboard/admin_app_version_form.html', {
        'form': form,
        'version_obj': version_obj,
    })


@login_required
def admin_app_version_delete(request, pk):
    if not require_admin(request.user):
        return HttpResponseForbidden()

    version_obj = get_object_or_404(AppVersion, pk=pk)
    if request.method == 'POST':
        label = str(version_obj)
        version_obj.delete()
        AdminLog.objects.create(action=f'تم حذف إصدار تطبيق: {label}')
        messages.success(request, 'تم حذف الإصدار بنجاح.')
        return redirect('dashboard:admin_app_versions')

    return render(request, 'dashboard/admin_app_version_delete_confirm.html', {
        'version_obj': version_obj,
    })


@login_required
def admin_toggle_registration(request):
    if not require_admin(request.user):
        return JsonResponse({'success': False}, status=403)
    
    event = Event.get_current()
    event.is_registration_closed = not event.is_registration_closed
    event.save(update_fields=['is_registration_closed'])
    
    status = 'مغلق' if event.is_registration_closed else 'مفتوح'
    AdminLog.objects.create(action=f'تم تغيير حالة التسجيل إلى: {status}')
    
    return JsonResponse({
        'success': True, 
        'is_closed': event.is_registration_closed,
        'message': f'تم {status} باب التسجيل بنجاح.'
    })


@login_required
def admin_toggle_existing_students_registration(request):
    if not require_admin(request.user):
        return JsonResponse({'success': False}, status=403)

    event = Event.get_current()
    event.allow_existing_students_registration = not event.allow_existing_students_registration
    event.save(update_fields=['allow_existing_students_registration'])

    status = 'مسموح' if event.allow_existing_students_registration else 'ممنوع'
    AdminLog.objects.create(action=f'تم تغيير حالة تسجيل الطلاب المسجلين سابقاً إلى: {status}')

    return JsonResponse({
        'success': True,
        'is_allowed': event.allow_existing_students_registration,
        'message': f'تم {status} تسجيل الطلاب المسجلين سابقاً.'
    })


@login_required
def admin_toggle_maintenance(request):
    if not require_admin(request.user):
        return JsonResponse({'success': False}, status=403)
    
    event = Event.get_current()
    event.is_maintenance_mode = not event.is_maintenance_mode
    event.save(update_fields=['is_maintenance_mode'])
    
    status = 'مفعل' if event.is_maintenance_mode else 'معطل'
    AdminLog.objects.create(action=f'تم تغيير حالة وضع الصيانة إلى: {status}')
    
    return JsonResponse({
        'success': True, 
        'is_maintenance': event.is_maintenance_mode,
        'message': f'تم {status} وضع الصيانة بنجاح.'
    })


def maintenance_page(request):
    event = Event.get_current()
    if not event.is_maintenance_mode:
        return redirect('dashboard:admin_dashboard')
    return render(request, 'dashboard/maintenance.html', {
        'facebook_url': event.maintenance_facebook_url
    })


@login_required
def admin_students_list(request):
    if not require_admin(request.user):
        return HttpResponseForbidden()
    
    event = Event.get_current()
    students_qs = Student.objects.select_related('group').all()
    q = request.GET.get('q') or ''
    group_id = request.GET.get('group') or ''
    if q:
        students_qs = students_qs.filter(Q(name__icontains=q) | Q(student_id__icontains=q))
    if group_id:
        students_qs = students_qs.filter(group_id=group_id)
    students = list(students_qs)
    for student in students:
        student.current_status = 'present' if student.checked_in else None
    
    # Show only groups for the current event in the filter dropdown
    groups = Group.objects.filter(event=event) if event else Group.objects.all()
    return render(request, 'dashboard/admin_students_list.html', {
        'students': students, 
        'groups': groups,
        'current_event': event
    })


@login_required
def admin_smart_search(request):
    if not require_admin(request.user):
        return HttpResponseForbidden()
    
    query = request.GET.get('q', '').strip()
    results = []
    
    if query:
        # البحث الشامل في الطلاب
        results = Student.objects.filter(
            Q(name__icontains=query) |
            Q(student_id__icontains=query) |
            Q(phone_number__icontains=query) |
            Q(email__icontains=query) |
            Q(school__icontains=query) |
            Q(education_admin__icontains=query) |
            Q(grade__icontains=query) |
            Q(group__name__icontains=query) |
            Q(group__code__icontains=query)
        ).select_related('group').distinct().order_by('name')
        
    return render(request, 'dashboard/admin_smart_search.html', {
        'query': query,
        'results': results,
        'current_event': Event.get_current()
    })


@login_required
def admin_current_event_students(request):
    if not require_admin(request.user):
        return HttpResponseForbidden()
    
    event = Event.get_current()
    if not event:
        messages.warning(request, 'لا توجد فعالية نشطة حالياً.')
        return redirect('dashboard:admin_dashboard')

    # Get students who have an APPROVED registration for the current event
    students_qs = Student.objects.filter(
        registrations__event=event,
        registrations__status=StudentRegistration.Status.APPROVED,
        registrations__removed_at__isnull=True,
    ).select_related('group').distinct()

    q = request.GET.get('q') or ''
    group_id = request.GET.get('group') or ''
    if q:
        students_qs = students_qs.filter(Q(name__icontains=q) | Q(student_id__icontains=q))
    if group_id:
        students_qs = students_qs.filter(group_id=group_id)
    
    # We want to know if they checked in for THIS event
    from students.models import StudentEventStats
    stats_map = {
        s.student_id: s.checked_in 
        for s in StudentEventStats.objects.filter(event=event)
    }

    students = list(students_qs)
    for student in students:
        # Use StudentEventStats if exists, fallback to student.checked_in (legacy)
        student.is_checked_in_this_event = stats_map.get(student.id, student.checked_in)
        
    groups = Group.objects.filter(event=event)
    return render(request, 'dashboard/admin_current_event_students.html', {
        'students': students, 
        'groups': groups,
        'event': event
    })


@login_required
def admin_student_create(request):
    if not require_admin(request.user):
        return HttpResponseForbidden()
    
    event = Event.get_current()
    groups = Group.objects.filter(event=event) if event else Group.objects.all()
    if request.method == 'POST':
        name = request.POST.get('name') or ''
        student_id = request.POST.get('student_id') or ''
        group_id = request.POST.get('group') or ''
        school = request.POST.get('school') or ''
        education_admin = request.POST.get('education_admin') or ''
        grade = request.POST.get('grade') or ''
        email = (request.POST.get('email') or '').strip()
        phone_number = (request.POST.get('phone_number') or '').strip()
        group = Group.objects.filter(id=group_id).first() if group_id else None
        if student_id and Student.objects.filter(student_id=student_id).exists():
            messages.error(request, 'رقم الطالب الذي أدخلته مستخدم بالفعل لطالب آخر.')
            return render(
                request,
                'dashboard/admin_student_form.html',
                {'groups': groups},
            )
        user = None
        password_plain = None
        if email:
            existing_user_with_email = User.objects.filter(email__iexact=email).exists()
            if existing_user_with_email:
                messages.error(request, 'هذا البريد الإلكتروني مستخدم بالفعل لحساب آخر، يرجى إدخال بريد مختلف.')
                return render(
                    request,
                    'dashboard/admin_student_form.html',
                    {'groups': groups},
                )
        if student_id:
            username = f'student_{student_id}'
            user, created = User.objects.get_or_create(
                username=username,
                defaults={'role': User.Roles.STUDENT, 'email': email},
            )
            if created:
                from django.utils.crypto import get_random_string

                password_plain = get_random_string(10)
                user.set_password(password_plain)
                user.save()
            else:
                if hasattr(user, 'student_profile'):
                    messages.error(
                        request,
                        'يوجد بالفعل طالب مرتبط بهذا حساب الدخول، لا يمكن ربط نفس الحساب بأكثر من طالب.',
                    )
                    return render(
                        request,
                        'dashboard/admin_student_form.html',
                        {'groups': groups},
                    )
        Student.objects.create(
            name=name,
            student_id=student_id,
            group=group,
            school=school,
            education_admin=education_admin,
            grade=grade,
            email=email,
            phone_number=phone_number,
            user=user,
            is_blacklisted=request.POST.get('is_blacklisted') == 'on',
            is_certificate_banned=request.POST.get('is_certificate_banned') == 'on',
        )
        AdminLog.objects.create(action=f'تم إضافة الطالب {name}')
        if password_plain and email:
            event = Event.get_current()
            whatsapp_block_text = ""
            whatsapp_block_html = ""
            if event.whatsapp_group_link:
                whatsapp_block_text = f"رابط مجموعة الواتساب للفعالية:\n{event.whatsapp_group_link}\n\n"
                whatsapp_block_html = f"""
                  <tr>
                    <td style="padding:12px 0 0 0;">
                      <table role="presentation" cellpadding="0" cellspacing="0" style="width:100%;border-radius:16px;background-color:#020617;border:1px solid #25d366;">
                        <tr>
                          <td style="padding:14px 16px;text-align:center;">
                            <p style="margin:0 0 10px;font-size:13px;color:#e5e7eb;font-weight:600;">
                              💬 مجموعة الواتساب الرسمية
                            </p>
                            <a href="{event.whatsapp_group_link}"
                               style="display:inline-block;padding:10px 18px;border-radius:999px;background-color:#25d366;color:#ffffff;font-size:13px;font-weight:700;text-decoration:none;">
                              الانضمام لمجموعة الواتساب
                            </a>
                          </td>
                        </tr>
                      </table>
                    </td>
                  </tr>
                """

            subject = 'بيانات حسابك في نظام Tech Day – EduTech Egypt'
            text_body = (
                f'مرحبًا {name},\n\n'
                f'تم إنشاء حساب لك على نظام متابعة الفعالية، ويمكنك استخدام البيانات التالية لتسجيل الدخول:\n\n'
                f'اسم المستخدم: {username}\n'
                f'كلمة المرور: {password_plain}\n\n'
                f'رابط تسجيل الدخول: https://td.edutech-egy.com/%D8%AD%D8%B3%D8%A7%D8%A8/%D8%AA%D8%B3%D8%AC%D9%8A%D9%84-%D8%AF%D8%AE%D9%88%D9%84/\n\n'
                f'{whatsapp_block_text}'
                f'يمكنك أيضًا استخدام رمز الـ QR المرفق لتسجيل حضورك.\n\n'
                f'تحياتنا،\n'
                f'EduTech Egypt System'
            )
            
            content_blocks = f"""
                <div class="td-email-box" style="padding:20px;border-radius:20px;background-color:#0f172a;border:1px solid #1e293b;margin-bottom:20px;">
                  <p class="td-email-text-main" style="margin:0 0 15px;font-size:14px;color:#e5e7eb;font-weight:700;">🔐 بيانات الدخول إلى حسابك</p>
                  <table role="presentation" cellpadding="0" cellspacing="0" width="100%">
                    <tr>
                      <td style="padding:8px 0;font-size:13px;color:#94a3b8;width:120px;">اسم المستخدم</td>
                      <td style="padding:8px 0;font-size:14px;font-family:monospace;color:#22d3ee;font-weight:700;">{username}</td>
                    </tr>
                    <tr>
                      <td style="padding:8px 0;font-size:13px;color:#94a3b8;">كلمة المرور</td>
                      <td style="padding:8px 0;font-size:14px;font-family:monospace;color:#f97316;font-weight:700;">{password_plain}</td>
                    </tr>
                  </table>
                </div>
                {whatsapp_block_html}
                <div class="td-email-box" style="padding:20px;border-radius:20px;background-color:#0f172a;border:1px solid #1e293b;text-align:center;">
                  <p class="td-email-text-main" style="margin:0 0 12px;font-size:14px;color:#e5e7eb;font-weight:700;">🎫 رمز الـ QR الخاص بحضورك</p>
                  <img src="https://api.qrserver.com/v1/create-qr-code/?size=180x180&data={student_id}" alt="QR Code" style="display:block;margin:0 auto;border-radius:12px;border:4px solid #ffffff;box-shadow:0 4px 15px rgba(0,0,0,0.3);">
                  <p class="td-email-text-muted" style="margin:12px 0 0;font-size:12px;color:#94a3b8;">برجاء إظهار هذا الكود عند بوابة الحضور لتسجيل دخولك.</p>
                </div>
            """
            footer_extra = f"""
                <a href="https://td.edutech-egy.com/%D8%AD%D8%B3%D8%A7%D8%A8/%D8%AA%D8%B3%D8%AC%D9%8A%D9%84-%D8%AF%D8%AE%D9%88%D9%84/" 
                   style="display:inline-block;padding:14px 32px;border-radius:999px;background:linear-gradient(135deg,#06b6d4,#6366f1);color:#ffffff;font-size:15px;font-weight:700;text-decoration:none;box-shadow:0 10px 25px rgba(99,102,241,0.4);">
                  🚀 دخول المنصة الآن
                </a>
            """
            
            html_body = get_styled_email_html(
                subject=subject,
                preview_text="بيانات دخولك لفعالية Tech Day + QR للحضور",
                title=f"👋 مرحبًا {name}",
                main_text="🎓 يسعدنا مشاركتك في فعالية <b>Tech Day</b> بالقليوبية.",
                content_blocks_html=content_blocks,
                footer_extra_html=footer_extra
            )
            
            message = EmailMultiAlternatives(
                subject,
                text_body,
                settings.DEFAULT_FROM_EMAIL,
                [email],
            )
            message.attach_alternative(html_body, 'text/html')
            send_email_async(message, 'إرسال بيانات دخول طالب جديد')
            messages.success(request, 'تم إضافة الطالب، سيتم إرسال رسالة الدخول إلى بريده الإلكتروني.')
        else:
            messages.success(request, 'تم إضافة الطالب بنجاح.')
        return redirect('dashboard:admin_students_list')
    return render(request, 'dashboard/admin_student_form.html', {'groups': groups})


@login_required
def admin_student_update(request, pk):
    if not require_admin(request.user):
        return HttpResponseForbidden()
    student = get_object_or_404(Student, pk=pk)
    
    event = Event.get_current()
    groups = Group.objects.filter(event=event) if event else Group.objects.all()
    if request.method == 'POST':
        old_group = student.group
        new_name = request.POST.get('name') or student.name
        new_student_id = request.POST.get('student_id') or student.student_id
        group_id = request.POST.get('group') or ''
        new_school = request.POST.get('school') or ''
        new_education_admin = request.POST.get('education_admin') or student.education_admin
        new_email = (request.POST.get('email') or '').strip()
        new_phone = (request.POST.get('phone_number') or '').strip()
        if new_email and User.objects.filter(email__iexact=new_email).exclude(id=student.user_id).exists():
            messages.error(request, 'هذا البريد الإلكتروني مستخدم بالفعل لحساب آخر، يرجى إدخال بريد مختلف.')
            return render(
                request,
                'dashboard/admin_student_form.html',
                {'student': student, 'groups': groups},
            )
        student.name = new_name
        student.student_id = new_student_id
        student.school = new_school
        student.education_admin = new_education_admin
        student.grade = request.POST.get('grade') or student.grade
        student.email = new_email or student.email
        student.phone_number = new_phone or student.phone_number
        student.is_blacklisted = request.POST.get('is_blacklisted') == 'on'
        student.is_certificate_banned = request.POST.get('is_certificate_banned') == 'on'
        group = Group.objects.filter(id=group_id).first() if group_id else None
        student.group = group
        student.save()
        AdminLog.objects.create(action=f'تم تحديث بيانات الطالب {student.name}')
        if group and group != old_group and student.email:
            subject = 'تحديث مجموعتك في فعالية Tech Day – EduTech Egypt'
            text_body = (
                f'مرحبًا {student.name},\n\n'
                f'تم تحديث مجموعتك في فعالية Tech Day – الفريق التقني بالقليوبية.\n\n'
                f'المجموعة الحالية: {group.name} ({group.code})\n\n'
                f'نتمنى لك تجربة مميزة وممتعة مع فريقك.\n\n'
                f'تحياتنا،\n'
                f'EduTech Egypt System'
            )
            
            content_blocks = f"""
                <div class="td-email-box" style="padding:20px;border-radius:20px;background-color:#0f172a;border:1px solid #1e293b;margin-bottom:20px;text-align:center;">
                  <p class="td-email-text-main" style="margin:0 0-10px;font-size:14px;color:#94a3b8;">مجموعتك الحالية:</p>
                  <div class="td-group-badge" style="display:inline-block;padding:8px 20px;border-radius:12px;background-color:#1e293b;color:#ffffff;font-size:18px;font-weight:800;border:1px solid {group.color};">
                    {group.name} ({group.code})
                  </div>
                </div>
            """
            
            html_body = get_styled_email_html(
                subject=subject,
                preview_text=f"تحديث مجموعتك: {group.name}",
                title="🔄 تحديث المجموعة",
                main_text=f"مرحبًا {student.name}، تم تغيير مجموعتك في الفعالية.",
                content_blocks_html=content_blocks
            )
            
            message = EmailMultiAlternatives(
                subject,
                text_body,
                settings.DEFAULT_FROM_EMAIL,
                [student.email],
            )
            message.attach_alternative(html_body, 'text/html')
            send_email_async(message, 'إرسال بيانات دخول طالب بعد التعديل')
        messages.success(request, 'تم تحديث بيانات الطالب')
        return redirect('dashboard:admin_students_list')
    return render(request, 'dashboard/admin_student_form.html', {'student': student, 'groups': groups})


@login_required
def admin_student_send_credentials(request, pk):
    if not require_admin(request.user):
        return HttpResponseForbidden()
    student = get_object_or_404(Student, pk=pk)
    if request.method != 'POST':
        return redirect('dashboard:admin_students_list')
    email = student.email or ''
    if not email:
        messages.error(request, 'لا يمكن إرسال بيانات الدخول لأن البريد الإلكتروني غير مسجل للطالب.')
        return redirect('dashboard:admin_students_list')
    user = student.user
    if not user and student.student_id:
        username = f'student_{student.student_id}'
        user, created = User.objects.get_or_create(
            username=username,
            defaults={'role': User.Roles.STUDENT, 'email': email},
        )
        if created:
            student.user = user
            student.save()
    if not user:
        messages.error(request, 'لا يمكن إرسال بيانات الدخول لأن حساب المستخدم غير متوفر.')
        return redirect('dashboard:admin_students_list')
    from django.utils.crypto import get_random_string

    password_plain = get_random_string(10)
    user.set_password(password_plain)
    user.email = email
    user.save()
    name = student.name
    username = user.username
    subject = 'تحديث بيانات حسابك في نظام Tech Day – EduTech Egypt'
    text_body = (
        f'مرحبًا {name},\n\n'
        f'تم تحديث بيانات الدخول الخاصة بحسابك على نظام متابعة فعالية Tech Day – الفريق التقني بالقليوبية.\n\n'
        f'يمكنك استخدام البيانات التالية لتسجيل الدخول:\n\n'
        f'اسم المستخدم: {username}\n'
        f'كلمة المرور الجديدة: {password_plain}\n\n'
        f'رابط تسجيل الدخول: https://td.edutech-egy.com/%D8%AD%D8%B3%D8%A7%D8%A8/%D8%AA%D8%B3%D8%AC%D9%8A%D9%84-%D8%AF%D8%AE%D9%88%D9%84/\n\n'
        f'يمكنك أيضًا استخدام رمز الـ QR المرفق في هذه الرسالة لتسجيل حضورك.\n\n'
        f'ننصحك بتغيير كلمة المرور بعد تسجيل الدخول للحفاظ على خصوصية حسابك.\n\n'
        f'في حال واجهت أي مشكلة في تسجيل الدخول يمكنك التواصل مع فريق الدعم.\n\n'
        f'تحياتنا،\n'
        f'EduTech Egypt System'
    )

    content_blocks = f"""
        <div class="td-email-box" style="padding:20px;border-radius:20px;background-color:#0f172a;border:1px solid #1e293b;margin-bottom:20px;">
          <p class="td-email-text-main" style="margin:0 0 15px;font-size:14px;color:#e5e7eb;font-weight:700;">🔐 بيانات الدخول المحدثة لحسابك</p>
          <table role="presentation" cellpadding="0" cellspacing="0" width="100%">
            <tr>
              <td style="padding:8px 0;font-size:13px;color:#94a3b8;width:120px;">اسم المستخدم</td>
              <td style="padding:8px 0;font-size:14px;font-family:monospace;color:#22d3ee;font-weight:700;">{username}</td>
            </tr>
            <tr>
              <td style="padding:8px 0;font-size:13px;color:#94a3b8;">كلمة المرور الجديدة</td>
              <td style="padding:8px 0;font-size:14px;font-family:monospace;color:#f97316;font-weight:700;">{password_plain}</td>
            </tr>
          </table>
        </div>
        <div class="td-email-box" style="padding:20px;border-radius:20px;background-color:#0f172a;border:1px solid #1e293b;text-align:center;">
          <p class="td-email-text-main" style="margin:0 0 12px;font-size:14px;color:#e5e7eb;font-weight:700;">🎫 رمز الـ QR الخاص بحضورك</p>
          <img src="https://api.qrserver.com/v1/create-qr-code/?size=180x180&data={student.student_id}" alt="QR Code" style="display:block;margin:0 auto;border-radius:12px;border:4px solid #ffffff;box-shadow:0 4px 15px rgba(0,0,0,0.3);">
          <p class="td-email-text-muted" style="margin:12px 0 0;font-size:12px;color:#94a3b8;">يمكنك استخدام هذا الكود لتسجيل حضورك في الفعالية بسرعة وسهولة.</p>
        </div>
    """
    footer_extra = f"""
        <a href="https://td.edutech-egy.com/%D8%AD%D8%B3%D8%A7%D8%A8/%D8%AA%D8%B3%D8%AC%D9%8A%D9%84-%D8%AF%D8%AE%D9%88%D9%84/" 
           style="display:inline-block;padding:14px 32px;border-radius:999px;background:linear-gradient(135deg,#06b6d4,#6366f1);color:#ffffff;font-size:15px;font-weight:700;text-decoration:none;box-shadow:0 10px 25px rgba(99,102,241,0.4);">
          🚀 دخول المنصة الآن
        </a>
        <p style="margin:20px 0 0;font-size:12px;color:#f97316;text-align:center;">
          ⚠️ ننصحك بتغيير كلمة المرور بعد تسجيل الدخول للحفاظ على خصوصية حسابك.
        </p>
    """

    html_body = get_styled_email_html(
        subject=subject,
        preview_text="تحديث بيانات دخولك لفعالية Tech Day + QR للحضور",
        title=f"🔄 تحديث بيانات الدخول",
        main_text=f"مرحبًا {name}، تم تحديث بيانات الدخول الخاصة بحسابك.",
        content_blocks_html=content_blocks,
        footer_extra_html=footer_extra
    )
    message = EmailMultiAlternatives(
        subject,
        text_body,
        settings.DEFAULT_FROM_EMAIL,
        [email],
    )
    message.attach_alternative(html_body, 'text/html')
    send_email_async(message, 'إرسال بيانات دخول طالب يدويًا')
    messages.success(request, 'تم إرسال بيانات الدخول الجديدة إلى بريد الطالب الإلكتروني (قد يستغرق الأمر لحظات).')
    return redirect('dashboard:admin_students_list')


@login_required
def admin_student_delete(request, pk):
    if not require_admin(request.user):
        return HttpResponseForbidden()
    student = get_object_or_404(Student, pk=pk)
    if request.method == 'POST':
        user = student.user
        user_to_delete = None
        if user and user.role == User.Roles.STUDENT and not user.is_staff and not user.is_superuser:
            user_to_delete = user
        student.delete()
        if user_to_delete:
            user_to_delete.delete()
        AdminLog.objects.create(action=f'تم حذف الطالب {student.name}')
        messages.success(request, 'تم حذف الطالب')
        return redirect('dashboard:admin_students_list')
    return render(request, 'dashboard/admin_student_delete_confirm.html', {'student': student})


@login_required
def admin_student_add_to_current_event(request, pk):
    if not require_admin(request.user):
        return HttpResponseForbidden()
    if request.method != 'POST':
        return HttpResponseForbidden()
    
    student = get_object_or_404(Student, pk=pk)
    event = Event.get_current()
    
    if student.is_registered_for_event(event):
        messages.warning(request, 'هذا الطالب مسجل بالفعل في الفعالية الحالية.')
    else:
        # Create a registration record
        StudentRegistration.objects.create(
            event=event,
            student=student,
            full_name_ar=student.name,
            email=student.email,
            phone_number=student.phone_number,
            school=student.school,
            education_admin=student.education_admin,
            grade=student.grade,
            status=StudentRegistration.Status.APPROVED,
            approved_by=request.user,
            approved_at=timezone.now()
        )
        AdminLog.objects.create(action=f'تم تسجيل الطالب {student.name} في الفعالية {event.name} يدوياً')
        messages.success(request, f'تم تسجيل الطالب في فعالية {event.name} بنجاح.')
        
    return redirect('dashboard:admin_students_list')


@login_required
def admin_student_remove_from_current_event(request, pk):
    if not require_admin(request.user):
        return HttpResponseForbidden()

    student = get_object_or_404(Student, pk=pk)
    event = Event.get_current()
    if not event:
        messages.error(request, 'لا توجد فعالية نشطة حالياً.')
        return redirect('dashboard:admin_dashboard')

    registration = (
        StudentRegistration.objects
        .filter(
            student=student,
            event=event,
            status=StudentRegistration.Status.APPROVED,
            removed_at__isnull=True,
        )
        .order_by('-created_at')
        .first()
    )

    if request.method == 'POST':
        if not registration:
            messages.warning(request, 'هذا الطالب غير مسجل في الفعالية الحالية.')
            return redirect('dashboard:admin_current_event_students')

        reason = (request.POST.get('reason') or '').strip()

        registration.removed_at = timezone.localtime()
        registration.removed_reason = reason
        registration.removed_by = request.user
        registration.save(update_fields=['removed_at', 'removed_reason', 'removed_by'])

        AdminLog.objects.create(
            action=f'تم إزالة الطالب {student.name} من الفعالية الحالية{": " + reason if reason else ""}',
        )

        email = (registration.email or student.email or '').strip()
        if email:
            subject = f'تم إلغاء تسجيلك في فعالية {event.name} – EduTech Egypt'
            reason_block_text = f'\n\nالسبب: {reason}\n' if reason else ''
            text_body = (
                f'مرحبًا {student.name},\n\n'
                f'نود إبلاغك بأنه تم إلغاء تسجيلك في الفعالية الحالية.\n'
                f'الفعالية: {event.name} – {event.location_name} ({event.year})\n'
                f'{reason_block_text}\n'
                f'في حال كان لديك أي استفسار يمكنك التواصل مع إدارة الفعالية.\n\n'
                f'تحياتنا،\n'
                f'EduTech Egypt System'
            )

            content_blocks = ''
            if reason:
                content_blocks = f"""
                    <div class="td-email-box" style="padding:20px;border-radius:20px;background-color:#0f172a;border:1px solid #1e293b;margin-bottom:20px;">
                      <p class="td-email-text-main" style="margin:0 0 10px;font-size:14px;color:#fca5a5;font-weight:700;">📝 سبب الإلغاء:</p>
                      <p class="td-email-text-main" style="margin:0;font-size:14px;color:#e5e7eb;line-height:1.6;">{reason}</p>
                    </div>
                """

            html_body = get_styled_email_html(
                subject=subject,
                preview_text='تم إلغاء تسجيلك في الفعالية الحالية',
                title='📌 تحديث مهم بخصوص تسجيلك',
                main_text=f'مرحبًا {student.name}، تم إلغاء تسجيلك في الفعالية الحالية.',
                content_blocks_html=content_blocks + f"""
                    <div class="td-email-box" style="padding:20px;border-radius:20px;background-color:#0f172a;border:1px solid #1e293b;">
                      <p class="td-email-text-main" style="margin:0;font-size:14px;color:#cbd5f5;">
                        الفعالية: <b>{event.name}</b> – {event.location_name} ({event.year})
                      </p>
                    </div>
                """,
            )

            message = EmailMultiAlternatives(
                subject,
                text_body,
                settings.DEFAULT_FROM_EMAIL,
                [email],
            )
            message.attach_alternative(html_body, 'text/html')
            send_email_async(message, 'إشعار إزالة طالب من الفعالية الحالية')
            messages.success(request, 'تمت إزالة الطالب من الفعالية وإرسال بريد له (إن وجد).')
        else:
            messages.success(request, 'تمت إزالة الطالب من الفعالية (لا يوجد بريد لإرسال رسالة).')

        return redirect('dashboard:admin_current_event_students')

    return render(request, 'dashboard/admin_remove_student_from_event_confirm.html', {
        'student': student,
        'event': event,
        'registration': registration,
    })


@login_required
def admin_registrations_list(request):
    if not require_admin(request.user):
        return HttpResponseForbidden()
    registrations = StudentRegistration.objects.select_related('student', 'approved_by').all()
    q = (request.GET.get('q') or '').strip()
    status = (request.GET.get('status') or '').strip().lower()
    if q:
        registrations = registrations.filter(
            Q(full_name_ar__icontains=q)
            | Q(email__icontains=q)
            | Q(phone_number__icontains=q)
            | Q(school__icontains=q)
        )
    if status in {StudentRegistration.Status.PENDING, StudentRegistration.Status.APPROVED, StudentRegistration.Status.REJECTED}:
        registrations = registrations.filter(status=status)

    registrations = registrations.order_by('-created_at')
    paginator = Paginator(registrations, 50)
    page_obj = paginator.get_page(request.GET.get('page') or 1)
    return render(
        request,
        'dashboard/admin_registrations_list.html',
        {
            'registrations': page_obj,
            'q': q,
            'status': status or 'all',
            'total_count': paginator.count,
        },
    )

@login_required
def admin_current_event_registrations(request):
    if not require_admin(request.user):
        return HttpResponseForbidden()
    event = Event.get_current()
    registrations = StudentRegistration.objects.select_related('student', 'approved_by').filter(event=event)
    status = (request.GET.get('status') or '').lower()
    if status == 'pending':
        registrations = registrations.filter(status=StudentRegistration.Status.PENDING)
    elif status == 'approved':
        registrations = registrations.filter(status=StudentRegistration.Status.APPROVED, removed_at__isnull=True)
    pending_count = StudentRegistration.objects.filter(event=event, status=StudentRegistration.Status.PENDING).count()
    approved_count = StudentRegistration.objects.filter(
        event=event, status=StudentRegistration.Status.APPROVED, removed_at__isnull=True
    ).count()
    registrations = registrations.order_by('-created_at')
    paginator = Paginator(registrations, 50)
    page_obj = paginator.get_page(request.GET.get('page') or 1)
    return render(
        request,
        'dashboard/admin_current_event_registrations.html',
        {
            'event': event,
            'registrations': page_obj,
            'status': status or 'all',
            'pending_count': pending_count,
            'approved_count': approved_count,
        },
    )

@login_required
def admin_registration_detail(request, pk):
    if not require_admin(request.user):
        return HttpResponseForbidden()
    registration = get_object_or_404(StudentRegistration, pk=pk)
    # Filter groups to only those belonging to the event of this registration
    groups = Group.objects.filter(event=registration.event)
    possible_student_matches = []
    possible_registration_matches = []
    if registration.status == StudentRegistration.Status.PENDING:
        email = (registration.email or '').strip()
        phone = (registration.phone_number or '').strip()
        name_ar = (registration.full_name_ar or '').strip()
        school = (registration.school or '').strip()
        possible_student_matches = list(
            Student.objects.filter(
                Q(email__iexact=email)
                | Q(phone_number=phone)
                | (Q(name__iexact=name_ar) & Q(school__iexact=school))
            )
            .distinct()
            .order_by('name')[:10]
        )
        possible_registration_matches = list(
            StudentRegistration.objects.filter(
                event=registration.event,
            )
            .filter(
                Q(email__iexact=email)
                | Q(phone_number=phone)
                | (Q(full_name_ar__iexact=name_ar) & Q(school__iexact=school))
            )
            .exclude(pk=registration.pk)
            .order_by('-created_at')[:10]
        )
    if request.method == 'POST':
        action = request.POST.get('action') or ''
        if action == 'approve' and registration.status == StudentRegistration.Status.PENDING:
            link_student_id = (request.POST.get('link_student_id') or '').strip()
            force_new = (request.POST.get('force_new') or '').strip() == '1'
            if (possible_student_matches or possible_registration_matches) and not force_new and not link_student_id:
                messages.warning(
                    request,
                    'تم العثور على تطابق محتمل. اختر "ربط بالطالب الموجود" أو فعّل خيار إنشاء طالب جديد رغم التطابق.',
                )
                return render(
                    request,
                    'dashboard/admin_registration_detail.html',
                    {
                        'registration': registration,
                        'groups': groups,
                        'possible_student_matches': possible_student_matches,
                        'possible_registration_matches': possible_registration_matches,
                    },
                )

            group_id = request.POST.get('group') or ''
            group = Group.objects.filter(id=group_id).first() if group_id else None

            student = None
            password_plain = None
            if link_student_id:
                student = Student.objects.filter(id=link_student_id).first()
                if not student:
                    messages.error(request, 'الطالب المختار غير موجود.')
                    return redirect('dashboard:admin_registration_detail', pk=registration.pk)
                if student.is_registered_for_event(registration.event):
                    registration.status = StudentRegistration.Status.REJECTED
                    registration.approved_by = request.user
                    registration.approved_at = timezone.localtime()
                    registration.save(update_fields=['status', 'approved_by', 'approved_at'])
                    AdminLog.objects.create(
                        action=f'تم رفض طلب تسجيل مكرر للطالب {registration.full_name_ar} (مسجل بالفعل في الفعالية).',
                    )
                    messages.info(request, 'تم رفض الطلب لأنه مكرر: الطالب مسجل بالفعل في الفعالية الحالية.')
                    next_reg = StudentRegistration.objects.filter(
                        event=registration.event,
                        status=StudentRegistration.Status.PENDING,
                    ).order_by('-created_at').first()
                    if next_reg:
                        return redirect('dashboard:admin_registration_detail', pk=next_reg.pk)
                    current_event = Event.get_current()
                    if current_event and registration.event_id == current_event.id:
                        return redirect('dashboard:admin_current_event_registrations')
                    return redirect('dashboard:admin_registrations_list')
                if group and not student.group:
                    student.group = group
                    student.save(update_fields=['group'])
            else:
                from django.utils.crypto import get_random_string

                student_id = None
                for _ in range(10):
                    candidate = get_random_string(6, allowed_chars='0123456789')
                    if not Student.objects.filter(student_id=candidate).exists():
                        student_id = candidate
                        break
                if not student_id:
                    messages.error(request, 'تعذّر توليد رقم فريد للطالب. حاول مرة أخرى.')
                    return redirect('dashboard:admin_registration_detail', pk=registration.pk)
                email = registration.email or ''
                user = None
                if email:
                    username = f'student_{student_id}'
                    if User.objects.filter(email__iexact=email).exists():
                        messages.error(
                            request,
                            'هذا البريد الإلكتروني مرتبط بالفعل بحساب آخر، لا يمكن إنشاء حساب جديد بنفس البريد.',
                        )
                        return redirect('dashboard:admin_registration_detail', pk=registration.pk)
                    user, created = User.objects.get_or_create(
                        username=username,
                        defaults={'role': User.Roles.STUDENT, 'email': email},
                    )
                    from django.utils.crypto import get_random_string as get_random_password

                    password_plain = get_random_password(10)
                    user.set_password(password_plain)
                    user.email = email
                    user.save()
                student = Student.objects.create(
                    name=registration.full_name_ar,
                    student_id=student_id,
                    group=group,
                    school=registration.school,
                    education_admin=registration.education_admin,
                    grade=registration.grade,
                    email=registration.email,
                    phone_number=registration.phone_number,
                    user=user,
                )
            registration.status = StudentRegistration.Status.APPROVED
            registration.student = student
            registration.approved_by = request.user
            registration.approved_at = timezone.localtime()
            registration.save()
            StudentEventStats.objects.get_or_create(student=student, event=registration.event)
            AdminLog.objects.create(
                action=f'تمت الموافقة على طلب تسجيل الطالب {registration.full_name_ar}',
            )
            email = (registration.email or '').strip()
            if link_student_id and email:
                if not (student.email or '').strip():
                    student.email = email
                    student.save(update_fields=['email'])
                send_registration_confirmation_email(student, registration.event)
                messages.success(
                    request,
                    'تمت الموافقة على الطلب وربطه بحساب طالب موجود. تم إرسال تفاصيل الفعالية إلى بريد الطالب.',
                )
                next_reg = StudentRegistration.objects.filter(
                    event=registration.event,
                    status=StudentRegistration.Status.PENDING,
                ).order_by('-created_at').first()
                if next_reg:
                    return redirect('dashboard:admin_registration_detail', pk=next_reg.pk)
                current_event = Event.get_current()
                if current_event and registration.event_id == current_event.id:
                    return redirect('dashboard:admin_current_event_registrations')
                return redirect('dashboard:admin_registrations_list')
            if link_student_id:
                messages.success(
                    request,
                    'تمت الموافقة على الطلب وربطه بحساب طالب موجود (لا يوجد بريد لإرسال رسالة).',
                )
                next_reg = StudentRegistration.objects.filter(
                    event=registration.event,
                    status=StudentRegistration.Status.PENDING,
                ).order_by('-created_at').first()
                if next_reg:
                    return redirect('dashboard:admin_registration_detail', pk=next_reg.pk)
                current_event = Event.get_current()
                if current_event and registration.event_id == current_event.id:
                    return redirect('dashboard:admin_current_event_registrations')
                return redirect('dashboard:admin_registrations_list')
            if password_plain and email:
                event = Event.get_current()
                whatsapp_block_text = ""
                whatsapp_block_html = ""
                if event.whatsapp_group_link:
                    whatsapp_block_text = f"رابط مجموعة الواتساب للفعالية:\n{event.whatsapp_group_link}\n\n"
                    whatsapp_block_html = f"""
                      <tr>
                        <td style="padding:12px 0 0 0;">
                          <table role="presentation" cellpadding="0" cellspacing="0" style="width:100%;border-radius:16px;background-color:#020617;border:1px solid #25d366;">
                            <tr>
                              <td style="padding:14px 16px;text-align:center;">
                                <p style="margin:0 0 10px;font-size:13px;color:#e5e7eb;font-weight:600;">
                                  💬 مجموعة الواتساب الرسمية
                                </p>
                                <a href="{event.whatsapp_group_link}"
                                   style="display:inline-block;padding:10px 18px;border-radius:999px;background-color:#25d366;color:#ffffff;font-size:13px;font-weight:700;text-decoration:none;">
                                  الانضمام لمجموعة الواتساب
                                </a>
                              </td>
                            </tr>
                          </table>
                        </td>
                      </tr>
                    """

                name = student.name
                username = user.username
                subject = 'تم قبول طلب تسجيلك في فعالية Tech Day – EduTech Egypt'
                text_body = (
                    f'مرحبًا {name},\n\n'
                    f'تم قبول طلب تسجيلك للمشاركة في فعالية Tech Day – الفريق التقني بالقليوبية.\n\n'
                    f'تم إنشاء حساب لك على نظام متابعة الفعالية، ويمكنك استخدام البيانات التالية لتسجيل الدخول:\n\n'
                    f'اسم المستخدم: {username}\n'
                    f'كلمة المرور: {password_plain}\n\n'
                    f'رابط تسجيل الدخول: https://td.edutech-egy.com/%D8%AD%D8%B3%D8%A7%D8%A8/%D8%AA%D8%B3%D8%AC%D9%8A%D9%84-%D8%AF%D8%AE%D9%88%D9%84/\n\n'
                    f'{whatsapp_block_text}'
                    f'يمكنك أيضًا استخدام رمز الـ QR المرفق في هذه الرسالة لتسجيل حضورك.\n\n'
                    f'ننصحك بتغيير كلمة المرور بعد أول تسجيل دخول للحفاظ على خصوصية حسابك.\n\n'
                    f'في حال واجهت أي مشكلة في تسجيل الدخول يمكنك التواصل مع فريق الدعم.\n\n'
                    f'تحياتنا،\n'
                    f'EduTech Egypt System'
                )
                content_blocks = f"""
                    <div class="td-email-box" style="padding:20px;border-radius:20px;background-color:#0f172a;border:1px solid #1e293b;margin-bottom:20px;">
                      <p class="td-email-text-main" style="margin:0 0 15px;font-size:14px;color:#e5e7eb;font-weight:700;">🔐 بيانات الدخول إلى حسابك</p>
                      <table role="presentation" cellpadding="0" cellspacing="0" width="100%">
                        <tr>
                          <td style="padding:8px 0;font-size:13px;color:#94a3b8;width:120px;">اسم المستخدم</td>
                          <td style="padding:8px 0;font-size:14px;font-family:monospace;color:#22d3ee;font-weight:700;">{username}</td>
                        </tr>
                        <tr>
                          <td style="padding:8px 0;font-size:13px;color:#94a3b8;">كلمة المرور</td>
                          <td style="padding:8px 0;font-size:14px;font-family:monospace;color:#f97316;font-weight:700;">{password_plain}</td>
                        </tr>
                      </table>
                    </div>
                    {whatsapp_block_html}
                    <div class="td-email-box" style="padding:20px;border-radius:20px;background-color:#0f172a;border:1px solid #1e293b;text-align:center;">
                      <p class="td-email-text-main" style="margin:0 0 12px;font-size:14px;color:#e5e7eb;font-weight:700;">🎫 رمز الـ QR الخاص بحضورك</p>
                      <img src="https://api.qrserver.com/v1/create-qr-code/?size=180x180&data={student_id}" alt="QR Code" style="display:block;margin:0 auto;border-radius:12px;border:4px solid #ffffff;box-shadow:0 4px 15px rgba(0,0,0,0.3);">
                      <p class="td-email-text-muted" style="margin:12px 0 0;font-size:12px;color:#94a3b8;">برجاء إظهار هذا الكود عند بوابة الحضور لتسجيل دخولك.</p>
                    </div>
                """
                footer_extra = f"""
                    <a href="https://td.edutech-egy.com/%D8%AD%D8%B3%D8%A7%D8%A8/%D8%AA%D8%B3%D8%AC%D9%8A%D9%84-%D8%AF%D8%AE%D9%88%D9%84/" 
                       style="display:inline-block;padding:14px 32px;border-radius:999px;background:linear-gradient(135deg,#06b6d4,#6366f1);color:#ffffff;font-size:15px;font-weight:700;text-decoration:none;box-shadow:0 10px 25px rgba(99,102,241,0.4);">
                      🚀 دخول المنصة الآن
                    </a>
                """
                
                html_body = get_styled_email_html(
                    subject=subject,
                    preview_text="تم قبول طلب تسجيلك في فعالية Tech Day + بيانات الدخول",
                    title=f"🎉 مرحبًا {name}",
                    main_text="تم قبول طلب تسجيلك للمشاركة في فعالية <b>Tech Day</b> بالقليوبية.",
                    content_blocks_html=content_blocks,
                    footer_extra_html=footer_extra
                )
                
                message = EmailMultiAlternatives(
                    subject,
                    text_body,
                    settings.DEFAULT_FROM_EMAIL,
                    [email],
                )
                message.attach_alternative(html_body, 'text/html')
                send_email_async(message, 'إرسال بيانات دخول لطالب تمت الموافقة على تسجيله')
                messages.success(
                    request,
                    'تمت الموافقة على طلب التسجيل، سيتم إرسال بيانات الدخول إلى بريد الطالب الإلكتروني.',
                )
            else:
                messages.success(
                    request,
                    'تمت الموافقة على طلب التسجيل وإنشاء حساب الطالب (بدون بريد إلكتروني).',
                )
            next_reg = StudentRegistration.objects.filter(
                event=registration.event,
                status=StudentRegistration.Status.PENDING,
            ).order_by('-created_at').first()
            if next_reg:
                return redirect('dashboard:admin_registration_detail', pk=next_reg.pk)
            current_event = Event.get_current()
            if current_event and registration.event_id == current_event.id:
                return redirect('dashboard:admin_current_event_registrations')
            return redirect('dashboard:admin_registrations_list')
        elif action == 'reject' and registration.status == StudentRegistration.Status.PENDING:
            rejection_reason = request.POST.get('rejection_reason', '').strip()
            registration.status = StudentRegistration.Status.REJECTED
            registration.approved_by = request.user
            registration.approved_at = timezone.localtime()
            registration.save()
            AdminLog.objects.create(
                action=f'تم رفض طلب تسجيل الطالب {registration.full_name_ar}',
            )
            email = registration.email or ''
            if email:
                name = registration.full_name_ar
                subject = 'نتيجة طلب التسجيل في فعالية Tech Day – EduTech Egypt'
                reason_block = ''
                if rejection_reason:
                    reason_block = (
                        f'سبب الرفض: {rejection_reason}\n\n'
                    )
                text_body = (
                    f'مرحبًا {name},\n\n'
                    f'نشكر لك اهتمامك بالمشاركة في فعالية Tech Day – الفريق التقني بالقليوبية.\n\n'
                    f'نود إبلاغك بأن طلب تسجيلك لم يتم قبوله في الوقت الحالي.\n\n'
                    f'{reason_block}'
                    f'هذا القرار لا يقلل من تقديرنا لك، ونتطلع لمشاركتك في فعاليات أخرى مستقبلًا بإذن الله.\n\n'
                    f'في حال كان لديك أي استفسار يمكنك التواصل مع فريق EDU-TECH في إدارتك التعليمية أو عبر منسق الفعالية.\n\n'
                    f'تحياتنا،\n'
                    f'EduTech Egypt System'
                )
                
                content_blocks = ""
                if rejection_reason:
                    content_blocks = f"""
                        <div class="td-email-box" style="padding:20px;border-radius:20px;background-color:#0f172a;border:1px solid #1e293b;margin-bottom:20px;">
                          <p class="td-email-text-main" style="margin:0 0 10px;font-size:14px;color:#fca5a5;font-weight:700;">📝 سبب عدم القبول:</p>
                          <p class="td-email-text-main" style="margin:0;font-size:14px;color:#e5e7eb;line-height:1.6;">{rejection_reason}</p>
                        </div>
                    """
                
                html_body = get_styled_email_html(
                    subject=subject,
                    preview_text="بخصوص طلب تسجيلك في فعالية Tech Day",
                    title="📝 نتيجة طلب التسجيل",
                    main_text=f"مرحبًا {name}، نشكرك على اهتمامك بالانضمام إلينا.",
                    content_blocks_html=content_blocks + f"""
                        <p class="td-email-text-main" style="margin:0;font-size:14px;color:#cbd5f5;text-align:center;">
                          نود إبلاغك بأن طلب تسجيلك لم يتم قبوله في الوقت الحالي.<br>
                          نتطلع لرؤيتك في فعالياتنا القادمة.
                        </p>
                    """
                )
                
                message = EmailMultiAlternatives(
                    subject,
                    text_body,
                    settings.DEFAULT_FROM_EMAIL,
                    [email],
                )
                message.attach_alternative(html_body, 'text/html')
                send_email_async(message, 'إرسال نتيجة رفض طلب تسجيل')
                messages.success(
                    request,
                    'تم رفض طلب التسجيل، سيتم إرسال بريد للطالب بنتيجة الطلب.',
                )
            else:
                messages.success(request, 'تم رفض طلب التسجيل.')
            next_reg = StudentRegistration.objects.filter(
                event=registration.event,
                status=StudentRegistration.Status.PENDING,
            ).order_by('-created_at').first()
            if next_reg:
                return redirect('dashboard:admin_registration_detail', pk=next_reg.pk)
            current_event = Event.get_current()
            if current_event and registration.event_id == current_event.id:
                return redirect('dashboard:admin_current_event_registrations')
            return redirect('dashboard:admin_registrations_list')
        else:
            messages.error(request, 'لا يمكن تنفيذ العملية على هذا الطلب.')
            current_event = Event.get_current()
            if current_event and registration.event_id == current_event.id:
                return redirect('dashboard:admin_current_event_registrations')
            return redirect('dashboard:admin_registrations_list')
    return render(
        request,
        'dashboard/admin_registration_detail.html',
        {
            'registration': registration,
            'groups': groups,
            'possible_student_matches': possible_student_matches,
            'possible_registration_matches': possible_registration_matches,
        },
    )


@login_required
def admin_student_transfer(request, pk):
    if not require_admin(request.user):
        return HttpResponseForbidden()
    student = get_object_or_404(Student, pk=pk)
    groups = Group.objects.all()
    if request.method == 'POST':
        group_id = request.POST.get('group') or ''
        group = Group.objects.filter(id=group_id).first() if group_id else None
        old_group = student.group
        student.group = group
        student.save()
        AdminLog.objects.create(action=f'تم نقل الطالب {student.name} من مجموعة {old_group} إلى {group}')
        if group and group != old_group and student.email:
            subject = 'تم نقلك إلى مجموعة جديدة في Tech Day – EduTech Egypt'
            text_body = (
                f'مرحبًا {student.name},\n\n'
                f'تم نقلك إلى مجموعة جديدة في فعالية Tech Day – الفريق التقني بالقليوبية.\n\n'
                f'المجموعة الحالية: {group.name} ({group.code})\n\n'
                f'نتمنى لك تجربة مميزة وممتعة مع فريقك.\n\n'
                f'تحياتنا،\n'
                f'EduTech Egypt System'
            )
            
            content_blocks = f"""
                <div class="td-email-box" style="padding:20px;border-radius:20px;background-color:#0f172a;border:1px solid #1e293b;margin-bottom:20px;text-align:center;">
                  <p class="td-email-text-main" style="margin:0 0 10px;font-size:14px;color:#94a3b8;">مجموعتك الجديدة:</p>
                  <div class="td-group-badge" style="display:inline-block;padding:8px 20px;border-radius:12px;background-color:#1e293b;color:#ffffff;font-size:18px;font-weight:800;border:1px solid {group.color};">
                    {group.name} ({group.code})
                  </div>
                </div>
            """
            
            html_body = get_styled_email_html(
                subject=subject,
                preview_text=f"تم نقلك لمجموعة جديدة: {group.name}",
                title="🔄 تغيير المجموعة",
                main_text=f"مرحبًا {student.name}، تم تحديث مجموعتك في الفعالية.",
                content_blocks_html=content_blocks
            )
            
            message = EmailMultiAlternatives(
                subject,
                text_body,
                settings.DEFAULT_FROM_EMAIL,
                [student.email],
            )
            message.attach_alternative(html_body, 'text/html')
            send_email_async(message, 'إشعار بنقل طالب إلى مجموعة جديدة')
        messages.success(request, 'تم نقل الطالب بنجاح')
        return redirect('dashboard:admin_students_list')
    return render(request, 'dashboard/admin_student_transfer.html', {'student': student, 'groups': groups})


@login_required
def admin_groups(request):
    if not require_admin(request.user):
        return HttpResponseForbidden()
    
    event = Event.get_current()
    show_all = request.GET.get('all') == 'true'
    
    if show_all:
        groups = Group.objects.all().order_by('-event__created_at', 'code')
    else:
        # عرض مجموعات الفعالية الحالية فقط افتراضياً
        groups = Group.objects.filter(event=event).order_by('code')
        
    return render(request, 'dashboard/admin_groups.html', {
        'groups': groups, 
        'event': event,
        'show_all': show_all
    })


@login_required
def admin_group_create(request):
    if not require_admin(request.user):
        return HttpResponseForbidden()
    event = Event.get_current()
    if request.method == 'POST':
        name = request.POST.get('name') or ''
        code = request.POST.get('code') or ''
        color = request.POST.get('color') or ''
        level = request.POST.get('level') or Group.Level.PRIMARY
        max_students = int(request.POST.get('max_students') or 0) or 25
        group = Group.objects.create(
            name=name, code=code, color=color, 
            max_students=max_students, level=level,
            event=event
        )
        AdminLog.objects.create(action=f'تم إنشاء المجموعة {group}')
        messages.success(request, 'تم إنشاء المجموعة بنجاح')
        return redirect('dashboard:admin_groups')
    return render(request, 'dashboard/admin_group_form.html')


@login_required
def admin_group_update(request, pk):
    if not require_admin(request.user):
        return HttpResponseForbidden()
    group = get_object_or_404(Group, pk=pk)
    if request.method == 'POST':
        group.name = request.POST.get('name') or group.name
        group.code = request.POST.get('code') or group.code
        group.color = request.POST.get('color') or group.color
        group.level = request.POST.get('level') or group.level
        group.max_students = int(request.POST.get('max_students') or group.max_students)
        group.save()
        AdminLog.objects.create(action=f'تم تعديل المجموعة {group}')
        messages.success(request, 'تم تعديل المجموعة')
        return redirect('dashboard:admin_groups')
    return render(request, 'dashboard/admin_group_form.html', {'group': group})


@login_required
def admin_groups_redistribute(request):
    if not require_admin(request.user):
        return HttpResponseForbidden()
    
    event = Event.get_current()
    if not event:
        messages.error(request, 'لا توجد فعالية نشطة لإعادة توزيع الطلاب عليها.')
        return redirect('dashboard:admin_groups')
    
    primary_groups = list(Group.objects.filter(event=event, level=Group.Level.PRIMARY))
    prep_sec_groups = list(Group.objects.filter(event=event, level=Group.Level.PREP_SEC))
    
    # تصنيف الطلاب المسجلين في الفعالية الحالية فقط
    all_students = list(Student.objects.filter(
        registrations__event=event,
        registrations__status=StudentRegistration.Status.APPROVED,
        registrations__removed_at__isnull=True,
    ).distinct())
    
    primary_students = []
    prep_sec_students = []
    other_students = []
    
    for s in all_students:
        grade = (s.grade or '').lower()
        if 'prim' in grade:
            primary_students.append(s)
        elif 'prep' in grade or 'sec' in grade:
            prep_sec_students.append(s)
        else:
            # طلاب بدون مرحلة محددة، سنعتبرهم مع الإعدادي والثانوي مؤقتاً أو حسب المتوفر
            other_students.append(s)
            
    # إذا لم توجد مجموعات من نوع معين، ندمج المجموعات المتاحة
    if not primary_groups and prep_sec_groups:
        prep_sec_students.extend(primary_students)
        primary_students = []
        primary_groups = prep_sec_groups
    elif not prep_sec_groups and primary_groups:
        primary_students.extend(prep_sec_students)
        prep_sec_students = []
        prep_sec_groups = primary_groups
    elif not primary_groups and not prep_sec_groups:
        messages.error(request, 'لا توجد مجموعات معرفة لهذه الفعالية.')
        return redirect('dashboard:admin_groups')
        
    # دمج الطلاب غير المصنفين مع المرحلة الأكبر
    prep_sec_students.extend(other_students)
    
    assignments = [] # قائمة بـ (student, group)
    
    # توزيع الابتدائي
    if primary_students and primary_groups:
        total_primary_capacity = sum(g.max_students for g in primary_groups)
        
        # إذا زاد عدد الطلاب الابتدائي عن سعة مجموعاتهم
        if len(primary_students) > total_primary_capacity:
            excess_count = len(primary_students) - total_primary_capacity
            # نقل الزيادة لطلاب الإعدادي والثانوي
            prep_sec_students.extend(primary_students[-excess_count:])
            primary_students = primary_students[:-excess_count]
            
        # توزيع الطلاب الابتدائي المتبقين بالتساوي على مجموعات الابتدائي
        for i, s in enumerate(primary_students):
            g = primary_groups[i % len(primary_groups)]
            assignments.append((s, g))
            
    # توزيع الإعدادي والثانوي (مع أي زيادة من الابتدائي)
    if prep_sec_students and prep_sec_groups:
        for i, s in enumerate(prep_sec_students):
            g = prep_sec_groups[i % len(prep_sec_groups)]
            assignments.append((s, g))
    elif prep_sec_students and primary_groups:
        # حالة طارئة: لا توجد مجموعات إعدادي/ثانوي، نوزعهم على الابتدائي
        for i, s in enumerate(prep_sec_students):
            g = primary_groups[i % len(primary_groups)]
            assignments.append((s, g))

    # تنفيذ التحديثات وإرسال الإيميلات
    with transaction.atomic():
        # تفريغ المجموعات الحالية للفعالية لضمان نظافة التوزيع
        Student.objects.filter(group__event=event).update(group=None)
        
        for student, new_group in assignments:
            old_group = student.group
            student.group = new_group
            student.save(update_fields=['group'])
            
            if new_group != old_group and student.email:
                # إرسال إيميل التحديث
                subject = 'تحديث مجموعتك بعد إعادة التوزيع في Tech Day – EduTech Egypt'
                
                content_blocks = f"""
                    <div class="td-email-box" style="padding:20px;border-radius:20px;background-color:#0f172a;border:1px solid #1e293b;margin-bottom:20px;text-align:center;">
                      <p class="td-email-text-main" style="margin:0 0 10px;font-size:14px;color:#94a3b8;">مجموعتك الحالية بعد التوزيع:</p>
                      <div class="td-group-badge" style="display:inline-block;padding:8px 20px;border-radius:12px;background-color:#1e293b;color:#ffffff;font-size:18px;font-weight:800;border:1px solid {new_group.color};">
                        {new_group.name} ({new_group.code})
                      </div>
                    </div>
                """
                
                html_body = get_styled_email_html(
                    subject=subject,
                    preview_text=f"إعادة توزيع المجموعات: مجموعتك هي {new_group.name}",
                    title="🔄 إعادة توزيع المجموعات",
                    main_text=f"مرحبًا {student.name}، تم تحديث بيانات مجموعتك في الفعالية.",
                    content_blocks_html=content_blocks
                )
                
                message = EmailMultiAlternatives(
                    subject,
                    f'مرحبًا {student.name},\n\nتم إعادة توزيعك للمجموعة: {new_group.name}\n\nنتمنى لك التوفيق.',
                    settings.DEFAULT_FROM_EMAIL,
                    [student.email],
                )
                message.attach_alternative(html_body, 'text/html')
                send_email_async(message, 'إشعار إعادة توزيع الطلاب على المجموعات')

    AdminLog.objects.create(action=f'تم إعادة توزيع {len(assignments)} طالب في فعالية {event.name} ({event.location_name})')
    messages.success(request, f'تم إعادة توزيع {len(assignments)} طالب في الفعالية الحالية بنجاح')
    return redirect('dashboard:admin_groups')


@login_required
def admin_send_location_email(request):
    if not require_admin(request.user):
        return HttpResponseForbidden()
    
    event = Event.get_current()
    if not event.location_name or not event.location_link:
        messages.error(request, 'يرجى ضبط مكان الفعالية ورابط الموقع أولاً في الإعدادات.')
        return redirect('dashboard:admin_event_settings')
    
    students = (
        Student.objects.filter(
            registrations__event=event,
            registrations__status=StudentRegistration.Status.APPROVED,
            registrations__removed_at__isnull=True,
            email__isnull=False,
        )
        .exclude(email='')
        .distinct()
    )
    if not students.exists():
        messages.error(request, 'لا يوجد طلاب في الفعالية الحالية لديهم بريد إلكتروني مسجل.')
        return redirect('dashboard:admin_dashboard')
    
    count = 0
    
    for student in students:
        subject = f'تأكيد حضور فعالية {event.name} - الموقع والموعد'
        name = student.name
        location = event.location_name
        map_link = event.location_link
        arrival_time = event.arrival_time_text or (event.start_datetime.strftime('%I:%M %p') if event.start_datetime else '8:00 AM')
        date_str = event.start_datetime.strftime('%Y-%m-%d') if event.start_datetime else 'يوم الفعالية'
        whatsapp_link = event.whatsapp_group_link

        text_body = (
            f'مرحبًا {name},\n\n'
            f'نؤكد حضورك لفعالية {event.name}.\n\n'
            f'📅 التاريخ: {date_str}\n'
            f'⏰ وقت الحضور المتوقع: {arrival_time}\n'
            f'📍 المكان: {location}\n'
            f'🗺️ رابط الموقع (Google Maps): {map_link}\n'
        )
        
        if whatsapp_link:
            text_body += f'💬 رابط مجموعة الواتساب: {whatsapp_link}\n'
            
        text_body += (
            f'\nنرجو الالتزام بموعد الحضور لضمان استفادة قصوى من كافة فقرات اليوم.\n\n'
            f'نتطلع لرؤيتك.\n\n'
            f'تحياتنا،\n'
            f'EduTech Egypt System'
        )

        whatsapp_section = ""
        if whatsapp_link:
            whatsapp_section = f"""
            <div class="td-email-box" style="padding:20px;border-radius:20px;background-color:#020617;border:1px solid #25d366;text-align:center;margin-top:20px;">
              <p style="margin:0 0 12px;font-size:14px;color:#e5e7eb;font-weight:600;">💬 انضم لمجموعة الواتساب الرسمية:</p>
              <a href="{whatsapp_link}" 
                 style="display:inline-block;padding:12px 24px;border-radius:999px;background-color:#25d366;color:#ffffff;font-size:14px;font-weight:700;text-decoration:none;">
                الانضمام لمجموعة الواتساب
              </a>
            </div>
            """

        content_blocks = f"""
            <div class="td-email-box" style="padding:24px;border-radius:24px;background-color:#0f172a;border:1px solid #1e293b;margin-bottom:20px;">
              <p class="td-email-text-main" style="margin:0 0 20px;font-size:15px;color:#22d3ee;font-weight:700;text-align:center;">📍 تفاصيل الحضور والموقع</p>
              
              <div style="background:#1e293b;border-radius:16px;padding:16px;margin-bottom:16px;">
                <table role="presentation" cellpadding="0" cellspacing="0" width="100%" style="direction:rtl;text-align:right;">
                  <tr>
                    <td style="padding:8px 0;font-size:13px;color:#94a3b8;width:100px;">📅 التاريخ</td>
                    <td style="padding:8px 0;font-size:14px;color:#e5e7eb;font-weight:600;">{date_str}</td>
                  </tr>
                  <tr>
                    <td style="padding:8px 0;font-size:13px;color:#94a3b8;">⏰ وقت الحضور</td>
                    <td style="padding:8px 0;font-size:14px;color:#f97316;font-weight:700;">{arrival_time}</td>
                  </tr>
                  <tr>
                    <td style="padding:8px 0;font-size:13px;color:#94a3b8;">📍 المكان</td>
                    <td style="padding:8px 0;font-size:14px;color:#e5e7eb;font-weight:600;">{location}</td>
                  </tr>
                </table>
              </div>

              <div style="text-align:center;">
                <a href="{map_link}" 
                   style="display:inline-block;padding:14px 32px;border-radius:999px;background-color:#22d3ee;color:#0f172a;font-size:15px;font-weight:700;text-decoration:none;box-shadow:0 10px 25px rgba(34,211,238,0.3);">
                  🗺️ فتح الموقع على الخريطة
                </a>
              </div>
            </div>

            {whatsapp_section}
            
            <p style="text-align:center;font-size:12px;color:#64748b;margin-top:24px;">
              نرجو الالتزام بموعد الحضور لضمان استفادة قصوى من كافة فقرات اليوم.
            </p>
        """
        
        html_body = get_styled_email_html(
            subject=subject,
            preview_text=f"تأكيد حضور فعالية {event.name}",
            title="✅ تأكيد حضور الفعالية",
            main_text=f"مرحبًا {name}، يسعدنا تأكيد حضورك لفعالية <b>{event.name}</b>.",
            content_blocks_html=content_blocks
        )

        message = EmailMultiAlternatives(
            subject,
            text_body,
            settings.DEFAULT_FROM_EMAIL,
            [student.email],
        )
        message.attach_alternative(html_body, 'text/html')
        send_email_async(message, 'إرسال إيميل تأكيد الموقع والموعد')
        count += 1

    AdminLog.objects.create(action=f'تم البدء في إرسال تأكيدات الحضور لـ {count} طالب')
    messages.success(request, f'جاري إرسال تأكيدات الحضور والموقع لطلاب الفعالية الحالية ({count}) في الخلفية.')
    return redirect('dashboard:admin_dashboard')


@login_required
def admin_workshops(request):
    if not require_admin(request.user):
        return HttpResponseForbidden()
    workshops = Workshop.objects.select_related('supervisor').all()
    return render(request, 'dashboard/admin_workshops.html', {'workshops': workshops})


@login_required
def admin_workshop_create(request):
    if not require_admin(request.user):
        return HttpResponseForbidden()
    supervisors = User.objects.filter(role=User.Roles.SUPERVISOR)
    current_event = Event.get_current()
    if request.method == 'POST':
        title = request.POST.get('title') or ''
        room = request.POST.get('room') or ''
        supervisor_id = request.POST.get('supervisor') or ''
        supervisor = supervisors.filter(id=supervisor_id).first() if supervisor_id else None
        status = request.POST.get('status') or 'upcoming'
        workshop = Workshop.objects.create(
            event=current_event,
            title=title,
            room=room,
            supervisor=supervisor,
            status=status,
        )
        AdminLog.objects.create(action=f'تم إنشاء الورشة {workshop.title}')
        if supervisor and supervisor.email:
            subject = 'تعيينك كمشرف ورشة في فعالية Tech Day – EduTech Egypt'
            name = supervisor.get_full_name() or supervisor.username
            text_body = (
                f'مرحبًا {name},\n\n'
                f'تم تعيينك كمشرف للورشة التالية في فعالية Tech Day – الفريق التقني بالقليوبية:\n\n'
                f'اسم الورشة: {workshop.title}\n'
                f'القاعة: {workshop.room}\n'
                f'حالة الورشة: {workshop.get_status_display()}\n\n'
                f'نرجو التواجد في القاعة قبل بدء الورشة بوقت كافٍ ومتابعة حضور الطلاب وتعليماتهم.\n\n'
                f'شكرًا لمشاركتك معنا.\n\n'
                f'تحياتنا،\n'
                f'EduTech Egypt System'
            )
            content_blocks = f"""
                <div class="td-email-box" style="padding:20px;border-radius:20px;background-color:#0f172a;border:1px solid #1e293b;margin-bottom:20px;">
                  <p class="td-email-text-main" style="margin:0 0 15px;font-size:14px;color:#22d3ee;font-weight:700;">📋 تفاصيل الورشة:</p>
                  <table role="presentation" cellpadding="0" cellspacing="0" width="100%">
                    <tr>
                      <td style="padding:8px 0;font-size:13px;color:#94a3b8;width:120px;">اسم الورشة</td>
                      <td style="padding:8px 0;font-size:14px;color:#e5e7eb;font-weight:700;">{workshop.title}</td>
                    </tr>
                    <tr>
                      <td style="padding:8px 0;font-size:13px;color:#94a3b8;">القاعة</td>
                      <td style="padding:8px 0;font-size:14px;color:#e5e7eb;">{workshop.room}</td>
                    </tr>
                    <tr>
                      <td style="padding:8px 0;font-size:13px;color:#94a3b8;">حالة الورشة</td>
                      <td style="padding:8px 0;font-size:14px;color:#e5e7eb;">{workshop.get_status_display()}</td>
                    </tr>
                  </table>
                </div>
                <div class="td-email-box" style="padding:20px;border-radius:20px;background-color:#0f172a;border:1px solid #1e293b;">
                  <p class="td-email-text-main" style="margin:0 0 12px;font-size:14px;color:#e5e7eb;font-weight:700;">📌 تعليمات المشرف:</p>
                  <ul style="margin:0;padding:0 20px;font-size:13px;color:#cbd5f5;line-height:1.6;">
                    <li>نرجو التواجد في القاعة المحددة قبل بدء الورشة بوقت كافٍ.</li>
                    <li>متابعة حضور الطلاب بدقة من خلال لوحة تحكم المشرف.</li>
                    <li>توجيه الطلاب ومساعدتهم خلال الأنشطة العملية.</li>
                  </ul>
                </div>
            """
            
            html_body = get_styled_email_html(
                subject=subject,
                preview_text=f"تم تعيينك مشرفاً لورشة: {workshop.title}",
                title="👨‍🏫 إشراف ورشة جديدة",
                main_text=f"مرحبًا {name}، يسعدنا انضمامك كأحد مشرفي ورش Tech Day.",
                content_blocks_html=content_blocks
            )
            
            message = EmailMultiAlternatives(
                subject,
                text_body,
                settings.DEFAULT_FROM_EMAIL,
                [supervisor.email],
            )
            message.attach_alternative(html_body, 'text/html')
            send_email_async(message, 'إشعار تعيين مشرف على ورشة')
        messages.success(request, 'تم إنشاء الورشة بنجاح')
        return redirect('dashboard:admin_workshops')
    return render(
        request,
        'dashboard/admin_workshop_form.html',
        {'supervisors': supervisors},
    )


@login_required
def admin_workshop_update(request, pk):
    if not require_admin(request.user):
        return HttpResponseForbidden()
    workshop = get_object_or_404(Workshop, pk=pk)
    supervisors = User.objects.filter(role=User.Roles.SUPERVISOR)
    current_event = Event.get_current()
    if request.method == 'POST':
        old_supervisor = workshop.supervisor
        workshop.title = request.POST.get('title') or workshop.title
        workshop.room = request.POST.get('room') or workshop.room
        supervisor_id = request.POST.get('supervisor') or ''
        new_supervisor = supervisors.filter(id=supervisor_id).first() if supervisor_id else None
        workshop.supervisor = new_supervisor
        workshop.status = request.POST.get('status') or workshop.status
        if not workshop.event_id and current_event:
            workshop.event = current_event
        workshop.save()
        AdminLog.objects.create(action=f'تم تعديل الورشة {workshop.title}')
        if new_supervisor and new_supervisor != old_supervisor and new_supervisor.email:
            subject = 'تعيينك أو تحديث إشرافك على ورشة في Tech Day – EduTech Egypt'
            name = new_supervisor.get_full_name() or new_supervisor.username
            text_body = (
                f'مرحبًا {name},\n\n'
                f'تم تعيينك أو تحديث إشرافك على الورشة التالية في فعالية Tech Day – الفريق التقني بالقليوبية:\n\n'
                f'اسم الورشة: {workshop.title}\n'
                f'القاعة: {workshop.room}\n'
                f'حالة الورشة: {workshop.get_status_display()}\n\n'
                f'نرجو التواجد في القاعة قبل بدء الورشة بوقت كافٍ ومتابعة حضور الطلاب وتعليماتهم.\n\n'
                f'شكرًا لمشاركتك معنا.\n\n'
                f'تحياتنا،\n'
                f'EduTech Egypt System'
            )
            content_blocks = f"""
                <div class="td-email-box" style="padding:20px;border-radius:20px;background-color:#0f172a;border:1px solid #1e293b;margin-bottom:20px;">
                  <p class="td-email-text-main" style="margin:0 0 15px;font-size:14px;color:#22d3ee;font-weight:700;">📋 تفاصيل الورشة المحدثة:</p>
                  <table role="presentation" cellpadding="0" cellspacing="0" width="100%">
                    <tr>
                      <td style="padding:8px 0;font-size:13px;color:#94a3b8;width:120px;">اسم الورشة</td>
                      <td style="padding:8px 0;font-size:14px;color:#e5e7eb;font-weight:700;">{workshop.title}</td>
                    </tr>
                    <tr>
                      <td style="padding:8px 0;font-size:13px;color:#94a3b8;">القاعة</td>
                      <td style="padding:8px 0;font-size:14px;color:#e5e7eb;">{workshop.room}</td>
                    </tr>
                    <tr>
                      <td style="padding:8px 0;font-size:13px;color:#94a3b8;">حالة الورشة</td>
                      <td style="padding:8px 0;font-size:14px;color:#e5e7eb;">{workshop.get_status_display()}</td>
                    </tr>
                  </table>
                </div>
                <div class="td-email-box" style="padding:20px;border-radius:20px;background-color:#0f172a;border:1px solid #1e293b;">
                  <p class="td-email-text-main" style="margin:0 0 12px;font-size:14px;color:#e5e7eb;font-weight:700;">📌 تعليمات المشرف:</p>
                  <ul style="margin:0;padding:0 20px;font-size:13px;color:#cbd5f5;line-height:1.6;">
                    <li>نرجو التواجد في القاعة المحددة قبل بدء الورشة بوقت كافٍ.</li>
                    <li>متابعة حضور الطلاب بدقة من خلال لوحة تحكم المشرف.</li>
                    <li>توجيه الطلاب ومساعدتهم خلال الأنشطة العملية.</li>
                  </ul>
                </div>
            """
            
            html_body = get_styled_email_html(
                subject=subject,
                preview_text=f"تحديث بيانات إشرافك لورشة: {workshop.title}",
                title="👨‍🏫 تحديث بيانات الإشراف",
                main_text=f"مرحبًا {name}، تم تحديث بيانات الورشة التي تشرف عليها.",
                content_blocks_html=content_blocks
            )
            
            message = EmailMultiAlternatives(
                subject,
                text_body,
                settings.DEFAULT_FROM_EMAIL,
                [new_supervisor.email],
            )
            message.attach_alternative(html_body, 'text/html')
            send_email_async(message, 'إشعار تحديث إشراف على ورشة')
        messages.success(request, 'تم تعديل الورشة')
        return redirect('dashboard:admin_workshops')
    return render(
        request,
        'dashboard/admin_workshop_form.html',
        {'workshop': workshop, 'supervisors': supervisors},
    )


@login_required
def admin_workshop_toggle_status(request, pk):
    if not require_admin(request.user):
        return HttpResponseForbidden()
    workshop = get_object_or_404(Workshop, pk=pk)
    if workshop.status == 'active':
        workshop.status = 'finished'
    else:
        workshop.status = 'active'
    workshop.save()
    AdminLog.objects.create(action=f'تم تغيير حالة الورشة {workshop.title} إلى {workshop.get_status_display()}')
    messages.success(request, 'تم تحديث حالة الورشة')
    return redirect('dashboard:admin_workshops')


@login_required
def admin_workshop_delete(request, pk):
    if not require_admin(request.user):
        return HttpResponseForbidden()
    workshop = get_object_or_404(Workshop, pk=pk)
    if request.method == 'POST':
        title = workshop.title
        workshop.delete()
        AdminLog.objects.create(action=f'تم حذف الورشة {title}')
        messages.success(request, 'تم حذف الورشة بنجاح')
        return redirect('dashboard:admin_workshops')
    return render(request, 'dashboard/admin_workshop_delete_confirm.html', {'workshop': workshop})


@login_required
def admin_supervisors(request):
    if not require_admin(request.user):
        return HttpResponseForbidden()
    supervisors = User.objects.filter(role=User.Roles.SUPERVISOR).annotate(
        workshop_count=Count('supervised_workshops')
    )
    return render(request, 'dashboard/admin_supervisors.html', {'supervisors': supervisors})


@login_required
def admin_supervisor_create(request):
    if not require_admin(request.user):
        return HttpResponseForbidden()
    if request.method == 'POST':
        first_name = request.POST.get('first_name') or ''
        last_name = request.POST.get('last_name') or ''
        email = (request.POST.get('email') or '').strip()
        if email and User.objects.filter(email__iexact=email).exists():
            messages.error(request, 'هذا البريد الإلكتروني مستخدم بالفعل لحساب آخر، يرجى إدخال بريد مختلف.')
            return render(request, 'dashboard/admin_supervisor_form.html')
        username_base = ''
        if email and '@' in email:
            username_base = email.split('@', 1)[0]
        if not username_base:
            username_base = 'supervisor'
        username = username_base
        suffix = 1
        while User.objects.filter(username=username).exists():
            suffix += 1
            username = f'{username_base}{suffix}'
        from django.utils.crypto import get_random_string

        password_plain = get_random_string(10)
        user = User(
            username=username,
            first_name=first_name,
            last_name=last_name,
            email=email,
            role=User.Roles.SUPERVISOR,
        )
        user.set_password(password_plain)
        user.save()
        AdminLog.objects.create(action=f'تم إنشاء مشرف جديد: {user.get_full_name() or user.username}')
        if email:
            subject = 'بيانات حسابك كمشرف ورشة في نظام Tech Day – EduTech Egypt'
            name = user.get_full_name() or username
            text_body = (
                f'مرحبًا {name},\n\n'
                f'تم إنشاء حساب لك كمشرف ورشة على نظام Tech Day – الفريق التقني بالقليوبية.\n\n'
                f'يمكنك استخدام البيانات التالية لتسجيل الدخول:\n\n'
                f'اسم المستخدم: {username}\n'
                f'كلمة المرور: {password_plain}\n\n'
                f'رابط تسجيل الدخول: https://td.edutech-egy.com/%D8%AD%D8%B3%D8%A7%D8%A8/%D8%AA%D8%B3%D8%AC%D9%8A%D9%84-%D8%AF%D8%AE%D9%88%D9%84/\n\n'
                f'ننصحك بتغيير كلمة المرور بعد أول تسجيل دخول للحفاظ على خصوصية حسابك.\n\n'
                f'في حال واجهت أي مشكلة في تسجيل الدخول يمكنك التواصل مع فريق الدعم.\n\n'
                f'تحياتنا،\n'
                f'EduTech Egypt System'
            )
            
            content_blocks = f"""
                <div class="td-email-box" style="padding:20px;border-radius:20px;background-color:#0f172a;border:1px solid #1e293b;margin-bottom:20px;">
                  <p class="td-email-text-main" style="margin:0 0 15px;font-size:14px;color:#e5e7eb;font-weight:700;">🔐 بيانات الدخول إلى حسابك</p>
                  <table role="presentation" cellpadding="0" cellspacing="0" width="100%">
                    <tr>
                      <td style="padding:8px 0;font-size:13px;color:#94a3b8;width:120px;">اسم المستخدم</td>
                      <td style="padding:8px 0;font-size:14px;font-family:monospace;color:#22d3ee;font-weight:700;">{username}</td>
                    </tr>
                    <tr>
                      <td style="padding:8px 0;font-size:13px;color:#94a3b8;">كلمة المرور</td>
                      <td style="padding:8px 0;font-size:14px;font-family:monospace;color:#f97316;font-weight:700;">{password_plain}</td>
                    </tr>
                  </table>
                </div>
                <div class="td-email-box" style="padding:20px;border-radius:20px;background-color:#0f172a;border:1px solid #1e293b;">
                  <p class="td-email-text-main" style="margin:0 0 12px;font-size:14px;color:#e5e7eb;font-weight:700;">🚀 البدء في العمل</p>
                  <ul style="margin:0;padding:0 20px;font-size:13px;color:#cbd5f5;line-height:1.6;">
                    <li>سجل دخولك للمنصة باستخدام البيانات أعلاه.</li>
                    <li>تأكد من الورش الموكلة إليك في جدول الورش.</li>
                    <li>قم بفتح صفحة الورشة عند البدء لتسجيل حضور الطلاب.</li>
                  </ul>
                </div>
            """
            
            footer_extra = f"""
                <a href="https://td.edutech-egy.com/%D8%AD%D8%B3%D8%A7%D8%A8/%D8%AA%D8%B3%D8%AC%D9%8A%D9%84-%D8%AF%D8%AE%D9%88%D9%84/" 
                   style="display:inline-block;padding:14px 32px;border-radius:999px;background:linear-gradient(135deg,#06b6d4,#6366f1);color:#ffffff;font-size:15px;font-weight:700;text-decoration:none;box-shadow:0 10px 25px rgba(99,102,241,0.4);">
                  🚀 دخول لوحة التحكم
                </a>
                <p style="margin:20px 0 0;font-size:12px;color:#f97316;text-align:center;">
                  ⚠️ ننصحك بتغيير كلمة المرور بعد أول تسجيل دخول للحفاظ على خصوصية حسابك.
                </p>
            """
            
            html_body = get_styled_email_html(
                subject=subject,
                preview_text="بيانات دخولك لمنصة Tech Day كمشرف ورشة",
                title=f"👋 مرحبًا {name}",
                main_text="👨‍🏫 تم إنشاء حساب لك كمشرف ورشة على نظام متابعة فعالية <b>Tech Day</b>.",
                content_blocks_html=content_blocks,
                footer_extra_html=footer_extra
            )
            
            message = EmailMultiAlternatives(
                subject,
                text_body,
                settings.DEFAULT_FROM_EMAIL,
                [email],
            )
            message.attach_alternative(html_body, 'text/html')
            send_email_async(message, 'إرسال بيانات دخول لمشرف جديد')
            messages.success(
                request,
                'تم إنشاء المشرف بنجاح، سيتم إرسال بيانات الدخول إلى بريده الإلكتروني.',
            )
        else:
            messages.success(request, 'تم إنشاء المشرف بنجاح.')
        return redirect('dashboard:admin_supervisors')
    return render(request, 'dashboard/admin_supervisor_form.html')


@login_required
def admin_supervisor_update(request, pk):
    if not require_admin(request.user):
        return HttpResponseForbidden()
    supervisor = get_object_or_404(User, pk=pk, role=User.Roles.SUPERVISOR)
    if request.method == 'POST':
        new_first_name = request.POST.get('first_name') or supervisor.first_name
        new_last_name = request.POST.get('last_name') or supervisor.last_name
        new_email = (request.POST.get('email') or '').strip()
        if new_email and User.objects.filter(email__iexact=new_email).exclude(id=supervisor.id).exists():
            messages.error(request, 'هذا البريد الإلكتروني مستخدم بالفعل لحساب آخر، يرجى إدخال بريد مختلف.')
            return render(request, 'dashboard/admin_supervisor_form.html', {'supervisor': supervisor})
        supervisor.first_name = new_first_name
        supervisor.last_name = new_last_name
        supervisor.email = new_email or supervisor.email
        supervisor.save()
        AdminLog.objects.create(action=f'تم تعديل بيانات المشرف {supervisor.get_full_name() or supervisor.username}')
        messages.success(request, 'تم تعديل بيانات المشرف')
        return redirect('dashboard:admin_supervisors')
    return render(request, 'dashboard/admin_supervisor_form.html', {'supervisor': supervisor})


@login_required
def admin_supervisor_delete(request, pk):
    if not require_admin(request.user):
        return HttpResponseForbidden()
    supervisor = get_object_or_404(User, pk=pk, role=User.Roles.SUPERVISOR)
    if request.method == 'POST':
        name = supervisor.get_full_name() or supervisor.username
        supervisor.delete()
        AdminLog.objects.create(action=f'تم حذف المشرف: {name}')
        messages.success(request, 'تم حذف المشرف بنجاح')
        return redirect('dashboard:admin_supervisors')
    return render(request, 'dashboard/admin_supervisor_delete_confirm.html', {'supervisor': supervisor})


@login_required
def admin_volunteers(request):
    if not require_admin(request.user):
        return HttpResponseForbidden()
    volunteers = User.objects.filter(role=User.Roles.VOLUNTEER)
    return render(request, 'dashboard/admin_volunteers.html', {'volunteers': volunteers})


@login_required
def admin_volunteer_create(request):
    if not require_admin(request.user):
        return HttpResponseForbidden()
    if request.method == 'POST':
        first_name = request.POST.get('first_name') or ''
        last_name = request.POST.get('last_name') or ''
        email = (request.POST.get('email') or '').strip()
        if email and User.objects.filter(email__iexact=email).exists():
            messages.error(request, 'هذا البريد الإلكتروني مستخدم بالفعل لحساب آخر، يرجى إدخال بريد مختلف.')
            return render(request, 'dashboard/admin_volunteer_form.html')
        username_base = ''
        if email and '@' in email:
            username_base = email.split('@', 1)[0]
        if not username_base:
            username_base = 'volunteer'
        username = username_base
        suffix = 1
        while User.objects.filter(username=username).exists():
            suffix += 1
            username = f'{username_base}{suffix}'
        from django.utils.crypto import get_random_string

        password_plain = get_random_string(10)
        user = User(
            username=username,
            first_name=first_name,
            last_name=last_name,
            email=email,
            role=User.Roles.VOLUNTEER,
        )
        user.set_password(password_plain)
        user.save()
        AdminLog.objects.create(action=f'تم إنشاء متطوع جديد: {user.get_full_name() or user.username}')
        if email:
            subject = 'بيانات حسابك كمتطوع في نظام Tech Day – EduTech Egypt'
            name = user.get_full_name() or username
            text_body = (
                f'مرحبًا {name},\n\n'
                f'تم إنشاء حساب لك كمتطوع على نظام Tech Day – الفريق التقني بالقليوبية.\n\n'
                f'يمكنك استخدام البيانات التالية لتسجيل الدخول:\n\n'
                f'اسم المستخدم: {username}\n'
                f'كلمة المرور: {password_plain}\n\n'
                f'رابط تسجيل الدخول: https://td.edutech-egy.com/%D8%AD%D8%B3%D8%A7%D8%A8/%D8%AA%D8%B3%D8%AC%D9%8A%D9%84-%D8%AF%D8%AE%D9%88%D9%84/\n\n'
                f'ننصحك بتغيير كلمة المرور بعد أول تسجيل دخول للحفاظ على خصوصية حسابك.\n\n'
                f'في حال واجهت أي مشكلة في تسجيل الدخول يمكنك التواصل مع فريق الدعم.\n\n'
                f'تحياتنا،\n'
                f'EduTech Egypt System'
            )
            
            content_blocks = f"""
                <div class="td-email-box" style="padding:20px;border-radius:20px;background-color:#0f172a;border:1px solid #1e293b;margin-bottom:20px;">
                  <p class="td-email-text-main" style="margin:0 0 15px;font-size:14px;color:#e5e7eb;font-weight:700;">🔐 بيانات الدخول إلى حسابك</p>
                  <table role="presentation" cellpadding="0" cellspacing="0" width="100%">
                    <tr>
                      <td style="padding:8px 0;font-size:13px;color:#94a3b8;width:120px;">اسم المستخدم</td>
                      <td style="padding:8px 0;font-size:14px;font-family:monospace;color:#22d3ee;font-weight:700;">{username}</td>
                    </tr>
                    <tr>
                      <td style="padding:8px 0;font-size:13px;color:#94a3b8;">كلمة المرور</td>
                      <td style="padding:8px 0;font-size:14px;font-family:monospace;color:#f97316;font-weight:700;">{password_plain}</td>
                    </tr>
                  </table>
                </div>
                <div class="td-email-box" style="padding:20px;border-radius:20px;background-color:#0f172a;border:1px solid #1e293b;">
                  <p class="td-email-text-main" style="margin:0 0 12px;font-size:14px;color:#e5e7eb;font-weight:700;">🚀 البدء في العمل</p>
                  <ul style="margin:0;padding:0 20px;font-size:13px;color:#cbd5f5;line-height:1.6;">
                    <li>سجل دخولك للمنصة باستخدام البيانات أعلاه.</li>
                    <li>تأكد من مهامك في لوحة تحكم المتطوع.</li>
                    <li>استخدم الماسح الضوئي لتسجيل حضور الطلاب في الورش.</li>
                  </ul>
                </div>
            """
            
            footer_extra = f"""
                <a href="https://td.edutech-egy.com/%D8%AD%D8%B3%D8%A7%D8%A8/%D8%AA%D8%B3%D8%AC%D9%8A%D9%84-%D8%AF%D8%AE%D9%88%D9%84/" 
                   style="display:inline-block;padding:14px 32px;border-radius:999px;background:linear-gradient(135deg,#06b6d4,#6366f1);color:#ffffff;font-size:15px;font-weight:700;text-decoration:none;box-shadow:0 10px 25px rgba(99,102,241,0.4);">
                  🚀 دخول لوحة التحكم
                </a>
                <p style="margin:20px 0 0;font-size:12px;color:#f97316;text-align:center;">
                  ⚠️ ننصحك بتغيير كلمة المرور بعد أول تسجيل دخول للحفاظ على خصوصية حسابك.
                </p>
            """
            
            html_body = get_styled_email_html(
                subject=subject,
                preview_text="بيانات دخولك لمنصة Tech Day كمتطوع",
                title=f"👋 مرحبًا {name}",
                main_text="🤝 تم إنشاء حساب لك كمتطوع على نظام متابعة فعالية <b>Tech Day</b>.",
                content_blocks_html=content_blocks,
                footer_extra_html=footer_extra
            )
            
            message = EmailMultiAlternatives(
                subject,
                text_body,
                settings.DEFAULT_FROM_EMAIL,
                [email],
            )
            message.attach_alternative(html_body, 'text/html')
            send_email_async(message, 'إرسال بيانات دخول لمتطوع جديد')
            messages.success(
                request,
                'تم إنشاء المتطوع بنجاح، سيتم إرسال بيانات الدخول إلى بريده الإلكتروني.',
            )
        else:
            messages.success(request, 'تم إنشاء المتطوع بنجاح.')
        return redirect('dashboard:admin_volunteers')
    return render(request, 'dashboard/admin_volunteer_form.html')


@login_required
def admin_volunteer_update(request, pk):
    if not require_admin(request.user):
        return HttpResponseForbidden()
    volunteer = get_object_or_404(User, pk=pk, role=User.Roles.VOLUNTEER)
    if request.method == 'POST':
        new_first_name = request.POST.get('first_name') or volunteer.first_name
        new_last_name = request.POST.get('last_name') or volunteer.last_name
        new_email = (request.POST.get('email') or '').strip()
        if new_email and User.objects.filter(email__iexact=new_email).exclude(id=volunteer.id).exists():
            messages.error(request, 'هذا البريد الإلكتروني مستخدم بالفعل لحساب آخر، يرجى إدخال بريد مختلف.')
            return render(request, 'dashboard/admin_volunteer_form.html', {'volunteer': volunteer})
        volunteer.first_name = new_first_name
        volunteer.last_name = new_last_name
        volunteer.email = new_email or volunteer.email
        volunteer.save()
        AdminLog.objects.create(
            action=f'تم تعديل بيانات المتطوع {volunteer.get_full_name() or volunteer.username}'
        )
        messages.success(request, 'تم تعديل بيانات المتطوع')
        return redirect('dashboard:admin_volunteers')
    return render(request, 'dashboard/admin_volunteer_form.html', {'volunteer': volunteer})


@login_required
def admin_volunteer_delete(request, pk):
    if not require_admin(request.user):
        return HttpResponseForbidden()
    volunteer = get_object_or_404(User, pk=pk, role=User.Roles.VOLUNTEER)
    if request.method == 'POST':
        name = volunteer.get_full_name() or volunteer.username
        volunteer.delete()
        AdminLog.objects.create(action=f'تم حذف المتطوع: {name}')
        messages.success(request, 'تم حذف المتطوع بنجاح')
        return redirect('dashboard:admin_volunteers')
    return render(request, 'dashboard/admin_volunteer_delete_confirm.html', {'volunteer': volunteer})


@login_required
def admin_schedule(request):
    if not require_admin(request.user):
        return HttpResponseForbidden()
    current_event = Event.get_current()
    sessions = (
        WorkshopSession.objects.select_related('workshop', 'group')
        .filter(group__event=current_event)
        .all()
    )
    periods = WorkshopSession.PERIOD_CHOICES
    groups = Group.objects.filter(event=current_event).order_by('code')
    return render(
        request,
        'dashboard/admin_schedule.html',
        {'sessions': sessions, 'periods': periods, 'groups': groups, 'event': current_event},
    )


@login_required
def admin_schedule_random_assign(request):
    if not require_admin(request.user):
        return HttpResponseForbidden()
    if request.method != 'POST':
        return redirect('dashboard:admin_schedule')
    current_event = Event.get_current()
    # جلب كافة الورش المتاحة في النظام لضمان توزيعها بشكل صحيح
    workshops = list(Workshop.objects.all())
    groups = list(Group.objects.filter(event=current_event))
    if not workshops or not groups:
        messages.error(request, 'يجب إضافة الورش والمجموعات أولًا قبل التوزيع التلقائي.')
        return redirect('dashboard:admin_schedule')
    WorkshopSession.objects.filter(group__event=current_event).delete()
    
    # الفترات التي يتم فيها توزيع ورش فعلياً
    workshop_periods = ['9:00-9:55', '10:05-10:40', '10:50-11:00', '11:10-11:45', '11:55-12:05', '12:05-12:50', '1:00-1:55']
    
    period_time_map = {
        '9:00-9:55': (time(9, 0), time(9, 55)),
        '10:05-10:40': (time(10, 5), time(10, 40)),
        '10:50-11:00': (time(10, 50), time(11, 0)),
        '11:10-11:45': (time(11, 10), time(11, 45)),
        '11:55-12:05': (time(11, 55), time(12, 5)),
        '12:05-12:50': (time(12, 5), time(12, 50)),
        '1:00-1:55': (time(13, 0), time(13, 55)),
    }
    
    groups_sorted = sorted(groups, key=lambda g: g.code)
    for period_index, period_value in enumerate(workshop_periods):
        start_time_value, end_time_value = period_time_map.get(period_value, (time(9, 0), time(10, 0)))
        for index, group in enumerate(groups_sorted):
            workshop = workshops[(index + period_index) % len(workshops)]
            WorkshopSession.objects.create(
                workshop=workshop,
                group=group,
                period=period_value,
                start_time=start_time_value,
                end_time=end_time_value,
            )
    
    AdminLog.objects.create(action='تم إنشاء جدول الورش تلقائيًا للفترات التعليمية')
    messages.success(request, 'تم توزيع المجموعات على الورش التعليمية تلقائيًا.')
    return redirect('dashboard:admin_schedule')


@login_required
def admin_session_create(request):
    if not require_admin(request.user):
        return HttpResponseForbidden()
    current_event = Event.get_current()
    # جلب كافة الورش المتاحة في النظام لضمان ظهورها جميعاً في القائمة
    workshops = Workshop.objects.all()
    groups = Group.objects.filter(event=current_event)
    session = WorkshopSession()
    group_id_initial = request.GET.get('group') or ''
    period_initial = request.GET.get('period') or ''
    if group_id_initial:
        session.group = groups.filter(id=group_id_initial).first()
    if period_initial:
        session.period = period_initial
    if request.method == 'POST':
        workshop_id = request.POST.get('workshop') or ''
        group_id = request.POST.get('group') or ''
        period = request.POST.get('period') or ''
        start_time = request.POST.get('start_time') or ''
        end_time = request.POST.get('end_time') or ''
        workshop = workshops.filter(id=workshop_id).first()
        group = groups.filter(id=group_id).first()
        session.workshop = workshop or session.workshop
        session.group = group or session.group
        session.period = period or session.period
        session.start_time = start_time or session.start_time
        session.end_time = end_time or session.end_time
        allowed_workshop_periods = {
            '9:00-9:55', '10:05-10:40', '10:50-11:00', '11:10-11:45', '11:55-12:05', '12:05-12:50', '1:00-1:55'
        }
        if period and period not in allowed_workshop_periods:
            messages.error(request, 'هذه الفترة ليست فترة ورش، لا يمكن إنشاء جلسة لها.')
        elif not workshop or not group or not period or not start_time or not end_time:
            messages.error(request, 'يرجى ملء جميع الحقول المطلوبة قبل حفظ الجلسة.')
        else:
            existing_session, created = WorkshopSession.objects.get_or_create(
                group=group,
                period=period,
                defaults={
                    'workshop': workshop,
                    'start_time': start_time,
                    'end_time': end_time,
                },
            )
            if not created:
                existing_session.workshop = workshop
                existing_session.start_time = start_time
                existing_session.end_time = end_time
                existing_session.save()
            AdminLog.objects.create(
                action='تم إنشاء جلسة جديدة في الجدول الزمني'
                if created
                else 'تم تعديل جلسة في الجدول الزمني'
            )
            messages.success(request, 'تم حفظ الجلسة في الجدول الزمني.')
            return redirect('dashboard:admin_schedule')
    period_choices = [
        (v, l)
        for v, l in WorkshopSession.PERIOD_CHOICES
        if v in {'9:00-9:55', '10:05-10:40', '10:50-11:00', '11:10-11:45', '11:55-12:05', '12:05-12:50', '1:00-1:55'}
    ]
    return render(
        request,
        'dashboard/admin_session_form.html',
        {'session': session, 'workshops': workshops, 'groups': groups, 'period_choices': period_choices},
    )


@login_required
def admin_session_delete(request, pk):
    if not require_admin(request.user):
        return HttpResponseForbidden()
    session = get_object_or_404(WorkshopSession, pk=pk)
    AdminLog.objects.create(action=f'تم حذف جلسة: {session}')
    session.delete()
    messages.success(request, 'تم حذف الجلسة من الجدول بنجاح.')
    return redirect('dashboard:admin_schedule')


@login_required
def admin_session_update(request, pk):
    if not require_admin(request.user):
        return HttpResponseForbidden()
    session = get_object_or_404(WorkshopSession, pk=pk)
    current_event = Event.get_current()
    # جلب كافة الورش المتاحة في النظام لضمان ظهورها جميعاً في القائمة
    workshops = Workshop.objects.all()
    groups = Group.objects.filter(event=current_event)
    if request.method == 'POST':
        workshop_id = request.POST.get('workshop') or ''
        group_id = request.POST.get('group') or ''
        period = request.POST.get('period') or session.period
        start_time = request.POST.get('start_time') or session.start_time
        end_time = request.POST.get('end_time') or session.end_time
        allowed_workshop_periods = {
            '9:00-9:55', '10:05-10:40', '10:50-11:00', '11:10-11:45', '11:55-12:05', '12:05-12:50', '1:00-1:55'
        }
        if period and period not in allowed_workshop_periods:
            messages.error(request, 'هذه الفترة ليست فترة ورش، لا يمكن إنشاء جلسة لها.')
            return redirect('dashboard:admin_session_update', pk=session.pk)
        session.workshop = workshops.filter(id=workshop_id).first() if workshop_id else session.workshop
        session.group = groups.filter(id=group_id).first() if group_id else session.group
        session.period = period
        session.start_time = start_time
        session.end_time = end_time
        session.save()
        AdminLog.objects.create(action='تم تعديل جلسة في الجدول الزمني')
        messages.success(request, 'تم تعديل الجلسة')
        return redirect('dashboard:admin_schedule')
    period_choices = [
        (v, l)
        for v, l in WorkshopSession.PERIOD_CHOICES
        if v in {'9:00-9:55', '10:05-10:40', '10:50-11:00', '11:10-11:45', '11:55-12:05', '12:05-12:50', '1:00-1:55'}
    ]
    return render(
        request,
        'dashboard/admin_session_form.html',
        {'session': session, 'workshops': workshops, 'groups': groups, 'period_choices': period_choices},
    )


@login_required
def admin_notifications(request):
    if not require_admin(request.user):
        return HttpResponseForbidden()
    notifications = Notification.objects.all()
    if request.method == 'POST':
        title = request.POST.get('title') or ''
        body = request.POST.get('body') or ''
        target = request.POST.get('target') or Notification.Target.ALL
        group_id = request.POST.get('group') or ''
        group = Group.objects.filter(id=group_id).first() if group_id else None
        notification = Notification.objects.create(
            title=title,
            body=body,
            target=target,
            group=group,
        )
        AdminLog.objects.create(action=f'تم إرسال تنبيه: {notification.title}')
        messages.success(request, 'تم إرسال التنبيه')
        return redirect('dashboard:admin_notifications')
    groups = Group.objects.all()
    return render(
        request,
        'dashboard/admin_notifications.html',
        {'notifications': notifications, 'groups': groups},
    )


@login_required
def admin_vip_invites(request):
    if not require_admin(request.user):
        return HttpResponseForbidden()
    event = Event.get_current()
    if not event.start_datetime or not event.location_name:
        messages.error(request, 'يرجى ضبط موعد ومكان الفعالية أولًا من إعدادات الفعالية قبل إرسال دعوات VIP.')
        invites = VIPInvite.objects.all()[:20]
        return render(
            request,
            'dashboard/admin_vip_invites.html',
            {'invites': invites},
        )
    if request.method == 'POST':
        name = (request.POST.get('name') or '').strip()
        email = (request.POST.get('email') or '').strip()
        title = (request.POST.get('title') or '').strip()
        vip_time = (request.POST.get('vip_time') or '').strip()
        if not name or not email or not title or not vip_time:
            messages.error(request, 'يرجى إدخال الاسم والبريد الإلكتروني والوظيفة أو الصفة ووقت حضور الضيف.')
            invites = VIPInvite.objects.all()[:20]
            return render(
                request,
                'dashboard/admin_vip_invites.html',
                {'invites': invites},
            )
        greeting_name = f'السيد الأستاذ {name}'
        event_dt = timezone.localtime(event.start_datetime)
        event_date = event_dt.strftime('%Y-%m-%d')
        weekday_index = event_dt.weekday()
        weekday_map = {
            0: 'الاثنين',
            1: 'الثلاثاء',
            2: 'الأربعاء',
            3: 'الخميس',
            4: 'الجمعة',
            5: 'السبت',
            6: 'الأحد',
        }
        event_day = weekday_map.get(weekday_index, '')
        subject = 'دعوة خاصة لحضور فعالية Tech Day – EduTech Egypt'
        text_body_lines = [
            greeting_name,
            title,
            '',
            'يسرنا دعوتكم لحضور فعالية Tech Day، ضمن أنشطة الفريق التقني بتعليم القليوبية.',
            '',
            'هذه الفعالية مخصصة لطلاب المدارس، وتهدف إلى تعريفهم بالمجالات التقنية الحديثة، مع تقديم أنشطة وفعاليات ممتعة ومفيدة تشجع الطلاب على الابتكار والتعلم التفاعلي في المجال التقني.',
            '',
            'حضوركم الكريم سيشكل دعمًا معنويًا كبيرًا ويترك أثرًا إيجابيًا مباشرًا على الطلاب ويحفزهم على المشاركة والتفاعل.',
            '',
            f'وذلك يوم {event_day} الموافق {event_date}',
            f'في تمام الساعة {vip_time}',
            f'ب{event.location_name}.',
        ]
        if event.location_link:
            text_body_lines.append(f'رابط موقع الفعالية على الخريطة: {event.location_link}')
        text_body_lines.extend(
            [
                '',
                'وتفضلوا بقبول فائق الاحترام والتقدير ،،،',
                '',
                'مقدمه لسيادتكم',
                'مدير مديرية التربية والتعليم بمحافظة القليوبية',
            ]
        )
        text_body = '\n'.join(text_body_lines)
        
        content_blocks = f"""
            <div class="td-email-box" style="padding:24px;border-radius:24px;background-color:#0f172a;border:1px solid #1e293b;">
              <p class="td-email-text-main" style="margin:0 0 16px;font-size:15px;color:#e5e7eb;line-height:1.8;">
                يسرنا دعوتكم لحضور فعالية <b>Tech Day</b>، ضمن أنشطة الفريق التقني بتعليم القليوبية.
              </p>
              <p class="td-email-text-main" style="margin:0 0 16px;font-size:14px;color:#cbd5f5;line-height:1.6;">
                تهدف هذه الفعالية إلى تعريف الطلاب بالمجالات التقنية الحديثة عبر أنشطة ممتعة ومفيدة تشجع على الابتكار والتعلم التفاعلي.
              </p>
              <div style="margin:20px 0;padding:16px;border-radius:16px;background-color:#020617;border:1px solid #1e293b;">
                <table role="presentation" cellpadding="0" cellspacing="0" width="100%">
                  <tr>
                    <td style="padding:6px 0;font-size:13px;color:#94a3b8;width:100px;">📅 الموعد</td>
                    <td style="padding:6px 0;font-size:14px;color:#e5e7eb;font-weight:700;">يوم {event_day} الموافق {event_date}</td>
                  </tr>
                  <tr>
                    <td style="padding:6px 0;font-size:13px;color:#94a3b8;">🕒 الوقت</td>
                    <td style="padding:6px 0;font-size:14px;color:#e5e7eb;font-weight:700;">في تمام الساعة {vip_time}</td>
                  </tr>
                  <tr>
                    <td style="padding:6px 0;font-size:13px;color:#94a3b8;">📍 المكان</td>
                    <td style="padding:6px 0;font-size:14px;color:#e5e7eb;font-weight:700;">{event.location_name}</td>
                  </tr>
                </table>
                {f'<p style="margin:12px 0 0;text-align:center;"><a href="{event.location_link}" target="_blank" style="display:inline-block;padding:8px 16px;border-radius:999px;background-color:#1e293b;color:#22d3ee;font-size:12px;text-decoration:none;border:1px solid #22d3ee;">📍 فتح الموقع على الخريطة</a></p>' if event.location_link else ''}
              </div>
              <p class="td-email-text-main" style="margin:0;font-size:14px;color:#cbd5f5;line-height:1.6;">
                حضوركم الكريم يشكل دعمًا معنويًا كبيرًا ويترك أثرًا إيجابيًا مباشرًا على الطلاب ويحفزهم على المشاركة والتفاعل.
              </p>
            </div>
        """
        
        footer_extra = f"""
            <div style="margin-top:20px;text-align:center;">
              <p style="margin:0;font-size:13px;color:#cbd5f5;">وتفضلوا بقبول فائق الاحترام والتقدير ،،،</p>
              <p style="margin:10px 0 0;font-size:12px;color:#94a3b8;">
                مقدمه لسيادتكم<br>
                <b style="color:#e5e7eb;">مدير مديرية التربية والتعليم بمحافظة القليوبية</b>
              </p>
            </div>
        """
        
        html_body = get_styled_email_html(
            subject=subject,
            preview_text=f"دعوة تشريف لحضور فعالية Tech Day - {name}",
            title="✨ دعوة خاصة",
            main_text=f"{greeting_name}<br><small style='color:#94a3b8;'>{title}</small>",
            content_blocks_html=content_blocks,
            footer_extra_html=footer_extra
        )
        
        invite = VIPInvite.objects.create(
            name=name,
            email=email,
            title=title,
            sent_at=timezone.now(),
        )
        message = EmailMultiAlternatives(
            subject,
            text_body,
            settings.DEFAULT_FROM_EMAIL,
            [email],
        )
        message.attach_alternative(html_body, 'text/html')
        send_email_async(message, f'إرسال دعوة VIP إلى {invite.name}')
        AdminLog.objects.create(action=f'تم جدولة إرسال دعوة VIP إلى {invite.name}')
        messages.success(request, 'تم جدولة إرسال الدعوة بنجاح، سيتم إرسالها إلى بريد الضيف.')
        return redirect('dashboard:admin_vip_invites')
    invites = VIPInvite.objects.all()[:20]
    return render(
        request,
        'dashboard/admin_vip_invites.html',
        {'invites': invites},
    )


@login_required
def admin_factory_reset(request):
    if not require_admin(request.user):
        return HttpResponseForbidden()
    if request.method != 'POST':
        return redirect('dashboard:admin_dashboard')
    with transaction.atomic():
        Attendance.objects.all().delete()
        WorkshopSession.objects.all().delete()
        Workshop.objects.all().delete()
        StudentRegistration.objects.all().delete()
        Student.objects.all().delete()
        Group.objects.all().delete()
        Notification.objects.all().delete()
        VIPInvite.objects.all().delete()
        AdminLog.objects.all().delete()
        event = Event.get_current()
        event.name = 'Tech Day'
        event.start_datetime = None
        event.location_name = ''
        event.location_link = ''
        event.is_finished = False
        event.save()
    AdminLog.objects.create(action='تم تنفيذ إعادة ضبط المصنع للنظام (حذف جميع البيانات التشغيلية)')
    messages.success(request, 'تم مسح جميع البيانات وإعادة ضبط النظام بنجاح.')
    return redirect('dashboard:admin_dashboard')

@login_required
def admin_reports(request):
    if not require_admin(request.user):
        return HttpResponseForbidden()
    
    event = Event.get_current()
    if not event:
        event = Event.objects.create(name="Tech Day", location_name="Main", year=2026, is_active=True)

    # الطلاب المسجلين والموافق عليهم في هذه الفعالية فقط
    current_students_ids = StudentRegistration.objects.filter(
        event=event, 
        status=StudentRegistration.Status.APPROVED,
        removed_at__isnull=True,
    ).values_list('student_id', flat=True)
    
    current_students_qs = Student.objects.filter(id__in=current_students_ids)
    
    total_students = current_students_qs.count()
    total_attendance = current_students_qs.filter(checked_in=True).count()
    
    education_admins = current_students_qs.annotate(
        admin_clean=Trim('education_admin')
    ).exclude(
        admin_clean__isnull=True
    ).exclude(
        admin_clean=''
    ).values_list(
        'admin_clean', flat=True
    ).distinct().order_by('admin_clean')
    
    # إحصائيات المجموعات حسب النقاط والحضور (لهذه الفعالية فقط)
    by_group = (
        Group.objects.filter(event=event).annotate(
            present_count=Count(
                'students',
                filter=Q(students__id__in=current_students_ids, students__checked_in=True),
                distinct=True,
            )
        )
        .values('name', 'code', 'present_count', 'points')
        .order_by('-points')
    )
    
    # إحصائيات الورش حسب الحضور والتقييم (لهذه الفعالية فقط)
    feedback_stats = WorkshopFeedback.objects.filter(workshop=OuterRef('pk')).values('workshop').annotate(
        c=Count('id'),
        a=Avg('rating')
    )
    
    by_workshop = (
        Workshop.objects.filter(event=event).annotate(
            present_count=Count(
                'sessions__attendance_records',
                filter=Q(sessions__attendance_records__status=Attendance.Status.PRESENT, sessions__attendance_records__student__id__in=current_students_ids),
                distinct=True,
            ),
            feedbacks_count=Subquery(feedback_stats.values('c')),
            avg_rating=Subquery(feedback_stats.values('a'))
        )
        .values('id', 'title', 'present_count', 'feedbacks_count', 'avg_rating')
        .order_by('-avg_rating')
    )
    
    # أفضل الطلاب حسب النقاط في هذه الفعالية
    top_students = current_students_qs.order_by('-points')[:10]
    
    most_attended = by_workshop.order_by('-present_count').first() if by_workshop else None
    
    # الطلاب المستحقون للشهادة (حضروا الفعالية الحالية فقط)
    eligible_count = (
        current_students_qs.filter(checked_in=True, email__isnull=False, is_certificate_banned=False)
        .exclude(email='')
        .count()
    )
    
    context = {
        'event': event,
        'total_students': total_students,
        'total_attendance': total_attendance,
        'eligible_count': eligible_count,
        'by_group': by_group,
        'by_workshop': by_workshop,
        'top_students': top_students,
        'most_attended': most_attended,
        'education_admins': education_admins,
    }
    return render(request, 'dashboard/admin_reports.html', context)

@login_required
def admin_statistics(request):
    if not require_admin(request.user):
        return HttpResponseForbidden()
    
    event = Event.get_current()
    if not event:
        event = Event.objects.create(name="Tech Day", location_name="Main", year=2026, is_active=True)

    # 1. إحصائيات عامة (للفعالية الحالية)
    # الطلاب المسجلين والموافق عليهم في هذه الفعالية
    current_registrations = StudentRegistration.objects.filter(
        event=event,
        status=StudentRegistration.Status.APPROVED,
        removed_at__isnull=True,
    )
    current_students_ids = current_registrations.values_list('student_id', flat=True)
    total_students = current_registrations.count()
    
    # الطلاب الذين حضروا بالفعل في هذه الفعالية
    total_checked_in = Student.objects.filter(id__in=current_students_ids, checked_in=True).count()
    attendance_rate = (total_checked_in / total_students * 100) if total_students > 0 else 0
    
    # 2. إحصائيات حسب الإدارة التعليمية (للمسجلين في الفعالية فقط)
    stats_by_admin = current_registrations.values('education_admin').annotate(
        count=Count('id'),
        checked_in_count=Count('id', filter=Q(student__id__in=current_students_ids, student__checked_in=True))
    ).order_by('-count')
    
    # 3. إحصائيات حسب السنة الدراسية (للمسجلين في الفعالية فقط)
    grade_map = {
        '4-prim': 'الرابع الابتدائي',
        '5-prim': 'الخامس الابتدائي',
        '6-prim': 'السادس الابتدائي',
        '1-prep': 'الأول الإعدادي',
        '2-prep': 'الثاني الإعدادي',
        '3-prep': 'الثالث الإعدادي',
        '1-sec': 'الأول الثانوي',
        '2-sec': 'الثاني الثانوي',
        '3-sec': 'الثالث الثانوي',
    }
    stats_by_grade_raw = current_registrations.values('grade').annotate(count=Count('id')).order_by('grade')
    stats_by_grade = []
    for item in stats_by_grade_raw:
        stats_by_grade.append({
            'name': grade_map.get(item['grade'], item['grade'] or 'غير محدد'),
            'count': item['count']
        })
    
    # 4. إحصائيات الحضور حسب المجموعات (المرتبطة بالفعالية الحالية)
    stats_by_group = Group.objects.filter(event=event).annotate(
        total=Count('students'),
        checked_in=Count('students', filter=Q(students__checked_in=True))
    ).order_by('-checked_in')
    
    # 5. توزيع النقاط (أعلى 5 طلاب في الفعالية الحالية)
    # الطلاب الذين لديهم نقاط في هذه الفعالية
    top_points_students = Student.objects.filter(
        registrations__event=event, 
        registrations__status=StudentRegistration.Status.APPROVED,
        registrations__removed_at__isnull=True,
    ).order_by('-points')[:5]
    
    # 6. إحصائيات طلبات التسجيل (لهذه الفعالية)
    reg_stats = StudentRegistration.objects.filter(event=event).values('status').annotate(count=Count('id'))
    
    context = {
        'total_students': total_students,
        'total_checked_in': total_checked_in,
        'attendance_rate': round(attendance_rate, 1),
        'stats_by_admin': stats_by_admin,
        'stats_by_grade': stats_by_grade,
        'stats_by_group': stats_by_group,
        'top_points_students': top_points_students,
        'reg_stats': reg_stats,
    }
    return render(request, 'dashboard/admin_statistics.html', context)

@login_required
def admin_volunteer_notes(request):
    """
    View for admins to see notes/feedback from both volunteers and supervisors.
    """
    if not require_admin(request.user):
        return HttpResponseForbidden()
    
    show_all = request.GET.get('all') == 'true'
    today = timezone.localdate()
    
    notes = VolunteerNote.objects.select_related('author').order_by('-created_at')
    
    if not show_all:
        notes = notes.filter(created_at__date=today)
        
    return render(
        request,
        'dashboard/admin_volunteer_notes.html',
        {
            'notes': notes,
            'today': today,
            'show_all': show_all,
        },
    )


@login_required
def admin_student_violations(request):
    if not require_admin(request.user):
        return HttpResponseForbidden()
    if request.method == 'POST':
        violation_id = request.POST.get('violation_id') or ''
        action = request.POST.get('action') or ''
        if violation_id and action:
            violation = get_object_or_404(StudentViolation.objects.select_related('student'), pk=violation_id)
            student = violation.student
            if action == 'blacklist' and student:
                student.is_blacklisted = True
                student.save()
                violation.status = StudentViolation.Status.RESOLVED
                violation.admin_action = 'إضافة الطالب إلى القائمة السوداء'
                violation.handled_by = request.user
                violation.handled_at = timezone.now()
                violation.save()
                AdminLog.objects.create(
                    action=f'إضافة الطالب {student.name} ({student.student_id}) إلى القائمة السوداء بناءً على مخالفة متطوع'
                )
                messages.success(request, 'تم إضافة الطالب إلى القائمة السوداء وتحديث حالة المخالفة.')
            elif action == 'resolve':
                violation.status = StudentViolation.Status.RESOLVED
                violation.admin_action = 'تمت مراجعة المخالفة بدون إضافة إلى القائمة السوداء'
                violation.handled_by = request.user
                violation.handled_at = timezone.now()
                violation.save()
                AdminLog.objects.create(
                    action=f'تمت مراجعة مخالفة الطالب {student.name} ({student.student_id}) بدون إضافة إلى القائمة السوداء'
                )
                messages.success(request, 'تم تعليم المخالفة كمُعالَجة.')
            elif action == 'ban_certificate' and student:
                student.is_certificate_banned = True
                student.save()
                violation.status = StudentViolation.Status.RESOLVED
                violation.admin_action = 'حرمان الطالب من الشهادة'
                violation.handled_by = request.user
                violation.handled_at = timezone.now()
                violation.save()
                AdminLog.objects.create(
                    action=f'حرمان الطالب {student.name} ({student.student_id}) من الشهادة بناءً على مخالفة متطوع'
                )
                messages.success(request, 'تم حرمان الطالب من الشهادة بنجاح.')
        return redirect('dashboard:admin_student_violations')
    violations = (
        StudentViolation.objects.select_related('student', 'reported_by')
        .order_by('-created_at')
    )
    return render(
        request,
        'dashboard/admin_student_violations.html',
        {
            'violations': violations,
        },
    )


@login_required
def admin_reports_export_csv(request):
    if not require_admin(request.user):
        return HttpResponseForbidden()
    
    event = Event.get_current()
    current_students_ids = StudentRegistration.objects.filter(
        event=event, 
        status=StudentRegistration.Status.APPROVED,
        removed_at__isnull=True,
    ).values_list('student_id', flat=True)

    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = 'attachment; filename="techday_report.csv"'
    lines = ['اسم المجموعة,الكود,عدد الحضور\n']
    by_group = (
        Group.objects.filter(event=event).annotate(
            present_count=Count(
                'students',
                filter=Q(students__id__in=current_students_ids, students__checked_in=True),
                distinct=True,
            )
        )
        .values('name', 'code', 'present_count')
        .order_by('code')
    )
    for item in by_group:
        lines.append(f"{item['name']},{item['code']},{item['present_count']}\n")
    response.write(''.join(lines))
    return response


@login_required
def admin_export_students_by_admin(request):
    if not require_admin(request.user):
        return HttpResponseForbidden()
    
    event = Event.get_current()
    education_admin = (request.GET.get('education_admin') or '').strip()
    if not education_admin:
        return redirect('dashboard:admin_reports')
    
    qs = Student.objects.filter(
        registrations__event=event,
        registrations__status=StudentRegistration.Status.APPROVED,
        registrations__removed_at__isnull=True,
        education_admin__iexact=education_admin
    ).select_related('group', 'user').distinct().order_by('name')

    response = HttpResponse(content_type='text/csv; charset=utf-8')
    filename = f"students_{education_admin}.csv".replace(' ', '_')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    response.write('\ufeff')
    writer = csv.writer(response)
    writer.writerow([
        'الاسم',
        'رقم الطالب',
        'المجموعة',
        'كود المجموعة',
        'المدرسة',
        'الإدارة التعليمية',
        'البريد الإلكتروني',
        'السنة الدراسية',
        'تم تسجيل حضور الفعالية',
        'وقت تسجيل الحضور',
        'نقاط',
        'في القائمة السوداء',
        'محروم من الشهادة',
        'تاريخ الإضافة',
        'اسم المستخدم المرتبط',
    ])
    for s in qs:
        group_name = s.group.name if s.group else ''
        group_code = s.group.code if s.group else ''
        username = s.user.username if s.user else ''
        writer.writerow([
            s.name or '',
            s.student_id or '',
            group_name,
            group_code,
            s.school or '',
            s.education_admin or '',
            s.email or '',
            s.grade or '',
            'نعم' if s.checked_in else 'لا',
            s.checked_in_at.strftime('%Y-%m-%d %H:%M') if s.checked_in_at else '',
            s.points,
            'نعم' if s.is_blacklisted else 'لا',
            'نعم' if s.is_certificate_banned else 'لا',
            s.created_at.strftime('%Y-%m-%d %H:%M') if s.created_at else '',
            username,
        ])
    return response


@login_required
def admin_export_students_by_admin_excel(request):
    if not require_admin(request.user):
        return HttpResponseForbidden()
    
    event = Event.get_current()
    education_admin = (request.GET.get('education_admin') or '').strip()
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    except Exception:
        messages.error(request, 'مكتبة openpyxl غير متوفرة. يرجى تثبيتها: pip install openpyxl')
        return redirect('dashboard:admin_reports')
    
    if not education_admin or education_admin == 'all':
        # Get admins who have students in the current event
        admins = Student.objects.filter(
            registrations__event=event,
            registrations__status=StudentRegistration.Status.APPROVED,
            registrations__removed_at__isnull=True,
        ).annotate(admin_clean=Trim('education_admin')).exclude(
            admin_clean__isnull=True
        ).exclude(
            admin_clean=''
        ).values_list('admin_clean', flat=True).distinct().order_by('admin_clean')
    else:
        admins = [education_admin]
    
    wb = Workbook()
    ws = wb.active
    ws.title = 'بيانات الطلاب'
    ws.sheet_view.rightToLeft = True
    header = [
        'الاسم',
        'رقم الطالب',
        'المجموعة',
        'كود المجموعة',
        'المدرسة',
        'الإدارة التعليمية',
        'البريد الإلكتروني',
        'السنة الدراسية',
        'تم تسجيل حضور الفعالية',
        'وقت تسجيل الحضور',
        'النقاط',
        'في القائمة السوداء',
        'محروم من الشهادة',
        'تاريخ الإضافة',
        'اسم المستخدم المرتبط',
    ]
    ws.append(['كشف الطلاب حسب الإدارة التعليمية'])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(header))
    ws['A1'].font = Font(bold=True, size=14, color='00FFFFFF')
    ws['A1'].alignment = Alignment(horizontal='center')
    ws['A1'].fill = PatternFill('solid', fgColor='00222D3A')
    thin = Side(style='thin', color='0094a3b8')
    ws.append(header)
    for cell in ws[2]:
        cell.font = Font(bold=True, color='00e5e7eb')
        cell.fill = PatternFill('solid', fgColor='001e293b')
        cell.alignment = Alignment(horizontal='center')
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for admin_name in admins:
        ws.append([f'الإدارة التعليمية: {admin_name}'])
        ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=len(header))
        row = ws.max_row
        ws[f'A{row}'].font = Font(bold=True, color='000b1220')
        ws[f'A{row}'].fill = PatternFill('solid', fgColor='000ea5e9')
        ws[f'A{row}'].alignment = Alignment(horizontal='right')
        
        students = Student.objects.filter(
            registrations__event=event,
            registrations__status=StudentRegistration.Status.APPROVED,
            registrations__removed_at__isnull=True,
            education_admin__iexact=admin_name
        ).select_related('group', 'user').distinct().order_by('name')
        
        for s in students:
            group_name = s.group.name if s.group else ''
            group_code = s.group.code if s.group else ''
            username = s.user.username if s.user else ''
            row_data = [
                s.name or '',
                s.student_id or '',
                group_name,
                group_code,
                s.school or '',
                s.education_admin or '',
                s.email or '',
                s.grade or '',
                'نعم' if s.checked_in else 'لا',
                s.checked_in_at.strftime('%Y-%m-%d %H:%M') if s.checked_in_at else '',
                s.points,
                'نعم' if s.is_blacklisted else 'لا',
                'نعم' if s.is_certificate_banned else 'لا',
                s.created_at.strftime('%Y-%m-%d %H:%M') if s.created_at else '',
                username,
            ]
            ws.append(row_data)
            for c in ws[ws.max_row]:
                c.alignment = Alignment(horizontal='right')
                c.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    widths = [22, 16, 18, 14, 22, 22, 26, 14, 18, 20, 10, 16, 18, 18, 22]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(
        buffer.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    safe_admin = (education_admin or 'all').replace(' ', '_')
    response['Content-Disposition'] = f'attachment; filename="students_by_admin_{safe_admin}.xlsx"'
    return response


@login_required
def admin_export_verification_pages(request):
    if not require_admin(request.user):
        return HttpResponseForbidden()
    if request.method != 'POST':
        return redirect('dashboard:admin_reports')
    buffer = io.BytesIO()
    from zipfile import ZipFile, ZIP_DEFLATED

    event = Event.get_current()
    # تصدير صفحات التحقق فقط للطلاب الذين سجلوا حضوراً وغير محظورين من الشهادة
    students = Student.objects.filter(
        checked_in=True, 
        is_certificate_banned=False
    ).exclude(student_id__isnull=True).exclude(student_id='').select_related('group')
    
    with ZipFile(buffer, 'w', ZIP_DEFLATED) as zf:
        for student in students:
            attendance_qs = Attendance.objects.filter(
                student=student,
                status=Attendance.Status.PRESENT,
            ).select_related('session__workshop')
            present_count = attendance_qs.count()
            attended_workshops = list(
                attendance_qs.order_by('session__start_time')
                .values_list('session__workshop__title', flat=True)
                .distinct()
            )
            
            # حساب الإجمالي والنسبة للتوافق مع الشكل الجديد لـ verify.html
            assigned_sessions_count = 0
            attendance_rate = 0
            if student.group:
                assigned_sessions_count = WorkshopSession.objects.filter(group=student.group).count()
            
            # إذا كان الإجمالي صفراً، نستخدم المرجع العام (أقصى عدد جلسات لمجموعة)
            if assigned_sessions_count == 0:
                from django.db.models import Count
                max_sessions = WorkshopSession.objects.values('group').annotate(c=Count('id')).order_by('-c').first()
                if max_sessions:
                    assigned_sessions_count = max_sessions['c']

            if present_count > assigned_sessions_count:
                assigned_sessions_count = present_count
            
            if assigned_sessions_count > 0:
                attendance_rate = int((present_count / assigned_sessions_count) * 100)

            html_string = render_to_string(
                'students/verify.html',
                {
                    'student': student,
                    'event': event,
                    'present_count': present_count,
                    'attended_workshops': attended_workshops,
                    'assigned_sessions_count': assigned_sessions_count,
                    'attendance_rate': attendance_rate,
                },
            )
            identifier = student.student_id or str(student.id)
            filename = f'{identifier}.html'
            zf.writestr(filename, html_string)
    timestamp = timezone.now().strftime('%Y%m%d-%H%M%S')
    zip_name = f'techday-verification-pages-{timestamp}.zip'
    buffer.seek(0)
    response = HttpResponse(buffer.read(), content_type='application/zip')
    response['Content-Disposition'] = f'attachment; filename="{zip_name}"'
    AdminLog.objects.create(
        action=f'تم إنشاء ملف مضغوط يحتوي على صفحات تحقق الطلاب ({students.count()} ملف HTML)',
    )
    return response


@login_required
def admin_support_requests(request):
    if not require_admin(request.user):
        return HttpResponseForbidden()
    
    status_filter = request.GET.get('status')
    category_filter = request.GET.get('category')
    
    requests_qs = StudentSupportRequest.objects.select_related('student').all()
    
    if status_filter:
        requests_qs = requests_qs.filter(status=status_filter)
    if category_filter:
        requests_qs = requests_qs.filter(category=category_filter)
        
    groups = Group.objects.all()
        
    return render(
        request,
        'dashboard/admin_support_requests.html',
        {
            'support_requests': requests_qs,
            'categories': StudentSupportRequest.Category.choices,
            'statuses': StudentSupportRequest.Status.choices,
            'groups': groups,
        },
    )


@login_required
def admin_support_reply(request, pk):
    if not require_admin(request.user):
        return HttpResponseForbidden()
    
    support_request = get_object_or_404(StudentSupportRequest, pk=pk)
    
    if request.method == 'POST':
        reply = request.POST.get('reply', '').strip()
        status = request.POST.get('status', support_request.status)
        
        if reply:
            support_request.admin_reply = reply
            support_request.status = status
            support_request.save()
            
            AdminLog.objects.create(
                action=f'الرد على طلب دعم الطالب {support_request.student.name} (ID: {support_request.id})'
            )
            messages.success(request, 'تم حفظ الرد وتحديث حالة الطلب بنجاح.')
        else:
            messages.error(request, 'يرجى كتابة نص الرد.')
            
    return redirect('dashboard:admin_support_requests')


@login_required
def admin_student_mark_all_present(request, pk):
    if not require_admin(request.user):
        return HttpResponseForbidden()
    
    student = get_object_or_404(Student, pk=pk)
    
    # 1. تحديث حالة الطالب العامة
    student.is_present = True
    student.save()
    
    # 2. تحضير جلسات الورش المرتبطة بمجموعة الطالب فقط
    count = 0
    if student.group:
        sessions = WorkshopSession.objects.filter(group=student.group)
        
        for session in sessions:
            attendance, created = Attendance.objects.get_or_create(
                student=student,
                session=session,
                defaults={
                    'status': Attendance.Status.PRESENT,
                    'scanned_at': timezone.now()
                }
            )
            if not created and attendance.status != Attendance.Status.PRESENT:
                attendance.status = Attendance.Status.PRESENT
                attendance.scanned_at = timezone.now()
                attendance.save()
            count += 1
    else:
        # إذا لم يكن الطالب في مجموعة، لا يمكن تحديد جلساته بدقة، لذا نكتفي بالتحضير العام
        pass
            
    AdminLog.objects.create(
        action=f'تحضير كامل للطالب {student.name} في الفعالية وجميع الجلسات ({count} جلسة)'
    )
    
    messages.success(request, f'تم تحضير الطالب {student.name} بنجاح في الفعالية وفي جميع الجلسات الـ {count}.')
    return redirect(request.META.get('HTTP_REFERER', 'dashboard:admin_support_requests'))


@login_required
def admin_backup_page(request):
    if not require_admin(request.user):
        return HttpResponseForbidden()
    return render(request, 'dashboard/admin_backup.html')


@login_required
def admin_backup_download(request):
    if not require_admin(request.user):
        return HttpResponseForbidden()
    if request.method != 'POST':
        return redirect('dashboard:admin_backup_page')
    
    backup_type = request.POST.get('type', 'json')
    
    if backup_type == 'db_file':
        # تنزيل ملف قاعدة البيانات مباشرة
        db_path = settings.DATABASES['default']['NAME']
        if os.path.exists(db_path):
            with open(db_path, 'rb') as f:
                response = HttpResponse(f.read(), content_type='application/x-sqlite3')
                timestamp = timezone.now().strftime('%Y%m%d-%H%M%S')
                response['Content-Disposition'] = f'attachment; filename="techday-physical-db-{timestamp}.sqlite3"'
                AdminLog.objects.create(action=f'تم تنزيل ملف قاعدة البيانات الفعلي (SQLite)')
                return response
        else:
            messages.error(request, 'ملف قاعدة البيانات غير موجود.')
            return redirect('dashboard:admin_backup_page')

    # الخيار الافتراضي: JSON
    buffer = io.StringIO()
    try:
        call_command(
            'dumpdata',
            natural_foreign=True,
            natural_primary=True,
            indent=2,
            stdout=buffer,
        )
        data = buffer.getvalue()
        
        timestamp = timezone.now().strftime('%Y%m%d-%H%M%S')
        filename = f'techday-backup-{timestamp}.json'
        
        response = HttpResponse(data, content_type='application/json; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        AdminLog.objects.create(action=f'تم إنشاء وتنزيل نسخة احتياطية بصيغة JSON ({filename})')
        return response
    except Exception as e:
        AdminLog.objects.create(action=f'فشل إنشاء نسخة احتياطية: {e}')
        messages.error(request, 'حدث خطأ أثناء إنشاء النسخة الاحتياطية.')
        return redirect('dashboard:admin_backup_page')
    finally:
        buffer.close()


@login_required
def admin_backup_restore(request):
    if not require_admin(request.user):
        return HttpResponseForbidden()
    
    if request.method == 'POST' and request.FILES.get('backup_file'):
        backup_file = request.FILES['backup_file']
        file_ext = os.path.splitext(backup_file.name)[1].lower()
        
        if file_ext not in ['.json', '.sqlite3']:
            messages.error(request, 'يرجى رفع ملف بصيغة JSON أو SQLite3 فقط.')
            return redirect('dashboard:admin_backup_page')
        
        try:
            # Save the uploaded file temporarily
            with tempfile.NamedTemporaryFile(delete=False, suffix=file_ext) as tmp:
                for chunk in backup_file.chunks():
                    tmp.write(chunk)
                tmp_path = tmp.name
            
            try:
                if file_ext == '.json':
                    # منطق استرجاع JSON (مع مسح البيانات وحماية الأدمن)
                    # 1. مسح المستخدمين (ما عدا الأدمن)
                    User.objects.exclude(Q(is_superuser=True) | Q(role='admin')).delete()
                    Attendance.objects.all().delete()
                    WorkshopFeedback.objects.all().delete()
                    WorkshopSession.objects.all().delete()
                    Workshop.objects.all().delete()
                    Student.objects.all().delete()
                    StudentRegistration.objects.all().delete()
                    Group.objects.all().delete()
                    VIPInvite.objects.all().delete()
                    VolunteerNote.objects.all().delete()
                    StudentViolation.objects.all().delete()
                    Notification.objects.all().delete()
                    FailedEmail.objects.all().delete()
                    AppVersion.objects.all().delete()
                    AdminLog.objects.all().delete()
                    Event.objects.all().delete()
                    Badge.objects.all().delete()
                    StudentBadge.objects.all().delete()
                    StudentEventStats.objects.all().delete()
                    SOSRequest.objects.all().delete()
                    BroadcastMessage.objects.all().delete()
                    StudentSupportRequest.objects.all().delete()
                    PublicFormAnswer.objects.all().delete()
                    PublicFormSubmission.objects.all().delete()
                    PublicFormField.objects.all().delete()
                    PublicForm.objects.all().delete()
                    
                    # معالجة ملف JSON لتحويل EventSettings إلى Event لضمان التوافق
                    with open(tmp_path, 'r', encoding='utf-8') as f:
                        data = json.load(f)
                    
                    modified = False
                    for entry in data:
                        if entry.get('model') == 'dashboard.eventsettings':
                            entry['model'] = 'dashboard.event'
                            # إضافة الحقول الجديدة الناقصة في الموديل الجديد
                            if 'fields' in entry:
                                if 'year' not in entry['fields']:
                                    entry['fields']['year'] = 2026
                                if 'execution_number' not in entry['fields']:
                                    entry['fields']['execution_number'] = 1
                                if 'is_active' not in entry['fields']:
                                    entry['fields']['is_active'] = True
                                if 'is_archived' not in entry['fields']:
                                    entry['fields']['is_archived'] = False
                                if 'created_at' not in entry['fields']:
                                    # إضافة تاريخ حالي لضمان عدم حدوث خطأ NOT NULL
                                    entry['fields']['created_at'] = timezone.now().isoformat()
                            modified = True
                    
                    if modified:
                        with open(tmp_path, 'w', encoding='utf-8') as f:
                            json.dump(data, f, ensure_ascii=False, indent=4)
                    
                    call_command('loaddata', tmp_path)
                    
                    # 2. ضمان وجود فعالية نشطة وربط البيانات بها
                    # أولاً: مسح أي فعاليات فارغة قد تكون أنشئت تلقائياً أثناء عملية الرفع (بواسطة الميدل وير)
                    Event.objects.annotate(
                        w_count=Count('workshops'),
                        g_count=Count('groups'),
                        r_count=Count('registrations')
                    ).filter(w_count=0, g_count=0, r_count=0, name="Tech Day").exclude(id__in=Event.objects.order_by('created_at').values_list('id', flat=True)[:1]).delete()

                    # ثانياً: إذا لم توجد فعالية نشطة ولكن توجد فعاليات مستوردة، نفعل أحدث واحدة بها بيانات
                    if not Event.objects.filter(is_active=True).exists():
                        latest_event = Event.objects.annotate(
                            data_count=Count('workshops') + Count('groups')
                        ).order_by('-data_count', '-created_at').first()
                        
                        if latest_event:
                            latest_event.is_active = True
                            latest_event.save(update_fields=['is_active'])
                    
                    current_event = Event.get_current()
                    if current_event:
                        # ربط أي بيانات يتيمة (بدون فعالية) بالفعالية الحالية
                        Workshop.objects.filter(event__isnull=True).update(event=current_event)
                        Group.objects.filter(event__isnull=True).update(event=current_event)
                        StudentRegistration.objects.filter(event__isnull=True).update(event=current_event)
                        AdminLog.objects.filter(event__isnull=True).update(event=current_event)
                        SOSRequest.objects.filter(event__isnull=True).update(event=current_event)
                        BroadcastMessage.objects.filter(event__isnull=True).update(event=current_event)
                        
                        # حالة خاصة: إذا تم استيراد مجموعات مرتبطة بفعالية قديمة (ID مختلف) 
                        # والسيستم حالياً لا يحتوي إلا على الفعالية الحالية، نقوم بنقلهم للفعالية الحالية
                        if Event.objects.count() == 1:
                            Workshop.objects.exclude(event=current_event).update(event=current_event)
                            Group.objects.exclude(event=current_event).update(event=current_event)
                            StudentRegistration.objects.exclude(event=current_event).update(event=current_event)
                    
                    AdminLog.objects.create(action=f'تم استرجاع البيانات بنجاح من ملف JSON: {backup_file.name}')
                    messages.success(request, 'تم استرجاع بيانات JSON بنجاح.')
                
                elif file_ext == '.sqlite3':
                    # منطق استرجاع ملف القاعدة الفعلي
                    db_path = settings.DATABASES['default']['NAME']
                    # إغلاق الاتصال الحالي لضمان سلامة الاستبدال
                    from django.db import connections
                    connections.close_all()
                    
                    # استبدال الملف
                    shutil.copy2(tmp_path, db_path)
                    messages.success(request, 'تم استبدال ملف قاعدة البيانات بنجاح. يرجى تسجيل الدخول مرة أخرى إذا لزم الأمر.')
                    return redirect('dashboard:admin_dashboard')

            except Exception as e:
                AdminLog.objects.create(action=f'فشل استرجاع البيانات من {backup_file.name}: {e}')
                messages.error(request, f'حدث خطأ أثناء استرجاع البيانات: {e}')
            finally:
                if os.path.exists(tmp_path):
                    os.remove(tmp_path)
                    
        except Exception as e:
            messages.error(request, f'فشل التعامل مع الملف: {e}')
            
    return redirect('dashboard:admin_backup_page')


@login_required
def admin_send_whatsapp_link(request):
    if not require_admin(request.user):
        return HttpResponseForbidden()
    if request.method != 'POST':
        return redirect('dashboard:admin_event_settings')
    event = Event.get_current()
    if not event.whatsapp_group_link:
        messages.error(request, 'يرجى إدخال رابط مجموعة الواتساب أولًا.')
        return redirect('dashboard:admin_event_settings')
    students = (
        Student.objects.filter(
            registrations__event=event,
            registrations__status=StudentRegistration.Status.APPROVED,
            registrations__removed_at__isnull=True,
            email__isnull=False,
        )
        .exclude(email='')
        .distinct()
    )
    if not students.exists():
        messages.error(request, 'لا يوجد طلاب في الفعالية الحالية لديهم بريد إلكتروني مسجل لإرسال الرابط لهم.')
        return redirect('dashboard:admin_event_settings')

    def _job():
        for student in students:
            subject = f'رابط مجموعة واتساب فعالية {event.name}'
            text_body = (
                f'مرحبًا {student.name},\n\n'
                f'ندعوك للانضمام لمجموعة الواتساب الخاصة بفعالية {event.name} لمتابعة آخر التحديثات والتعليمات:\n\n'
                f'{event.whatsapp_group_link}\n\n'
                f'نتطلع لرؤيتك قريبًا.\n\n'
                f'تحياتنا،\n'
                f'EduTech Egypt System'
            )
            
            content_blocks = f"""
                <div class="td-email-box" style="padding:24px;border-radius:24px;background-color:#0f172a;border:1px solid #1e293b;text-align:center;">
                  <p class="td-email-text-main" style="margin:0 0 16px;font-size:15px;color:#e5e7eb;line-height:1.6;">
                    يرجى الانضمام للمجموعة الرسمية لمتابعة آخر التحديثات والتعليمات الخاصة بالفعالية:
                  </p>
                  <div style="margin:20px 0;">
                    <a href="{event.whatsapp_group_link}" 
                       style="display:inline-block;padding:14px 32px;border-radius:999px;background-color:#25d366;color:#ffffff;font-size:16px;font-weight:700;text-decoration:none;box-shadow:0 10px 25px rgba(37,211,102,0.4);">
                      🟢 الانضمام لمجموعة الواتساب
                    </a>
                  </div>
                  <p class="td-email-text-muted" style="margin:16px 0 0;font-size:12px;color:#94a3b8;">
                    إذا لم يعمل الزر، يمكنك استخدام الرابط المباشر:<br>
                    <a href="{event.whatsapp_group_link}" style="color:#22d3ee;text-decoration:none;">{event.whatsapp_group_link}</a>
                  </p>
                </div>
            """
            
            html_body = get_styled_email_html(
                subject=subject,
                preview_text=f"دعوة للانضمام لمجموعة واتساب فعالية {event.name}",
                title="💬 مجموعة الواتساب الرسمية",
                main_text=f"مرحبًا {student.name}، ندعوك للبقاء على تواصل معنا.",
                content_blocks_html=content_blocks
            )
            
            message = EmailMultiAlternatives(
                subject,
                text_body,
                settings.DEFAULT_FROM_EMAIL,
                [student.email],
            )
            message.attach_alternative(html_body, 'text/html')
            send_email_async(message, 'إرسال رابط مجموعة الواتساب')
        AdminLog.objects.create(action=f'تم إرسال رابط الواتساب لطلاب الفعالية الحالية لعدد {students.count()} طالب')

    threading.Thread(target=_job, daemon=True).start()
    messages.success(request, f'تم بدء عملية إرسال رابط الواتساب لطلاب الفعالية الحالية لـ {students.count()} طالب في الخلفية.')
    return redirect('dashboard:admin_event_settings')


@login_required
def admin_send_event_instructions(request):
    if not require_admin(request.user):
        return HttpResponseForbidden()
    if request.method != 'POST':
        return redirect('dashboard:admin_event_settings')
    event = Event.get_current()
    if not event.event_instructions:
        messages.error(request, 'يرجى إدخال تعليمات الفعالية أولاً.')
        return redirect('dashboard:admin_event_settings')
    students = (
        Student.objects.filter(
            registrations__event=event,
            registrations__status=StudentRegistration.Status.APPROVED,
            registrations__removed_at__isnull=True,
            email__isnull=False,
        )
        .exclude(email='')
        .distinct()
    )
    if not students.exists():
        messages.error(request, 'لا يوجد طلاب في الفعالية الحالية لديهم بريد إلكتروني مسجل لإرسال التعليمات لهم.')
        return redirect('dashboard:admin_event_settings')

    def _job():
        # Convert newlines to <br> for HTML email
        instructions_html = event.event_instructions.replace('\n', '<br>')
        for student in students:
            subject = f'تعليمات هامة لفعالية {event.name}'
            text_body = (
                f'مرحبًا {student.name},\n\n'
                f'يرجى قراءة التعليمات التالية الخاصة بفعالية {event.name}:\n\n'
                f'{event.event_instructions}\n\n'
                f'نتمنى لك يوماً سعيداً ومفيداً.\n\n'
                f'تحياتنا،\n'
                f'EduTech Egypt System'
            )
            
            content_blocks = f"""
                <div class="td-email-box" style="padding:24px;border-radius:24px;background-color:#0f172a;border:1px solid #1e293b;text-align:right;direction:rtl;">
                  <h3 style="margin:0 0 16px;color:#22d3ee;font-size:18px;font-weight:800;">📝 تعليمات الحضور</h3>
                  <p class="td-email-text-main" style="margin:0;font-size:15px;color:#e5e7eb;line-height:1.8;white-space:pre-line;">
                    {event.event_instructions}
                  </p>
                </div>
                <div style="margin-top:20px;text-align:center;">
                   <p style="color:#94a3b8;font-size:13px;">يرجى الالتزام بكافة التعليمات لضمان أفضل تجربة تعليمية.</p>
                </div>
            """
            
            html_body = get_styled_email_html(
                subject=subject,
                preview_text=f"تعليمات هامة لفعالية {event.name}",
                title="📋 تعليمات وإرشادات الفعالية",
                main_text=f"مرحبًا {student.name}، نرجو قراءة التعليمات التالية بعناية.",
                content_blocks_html=content_blocks
            )
            
            message = EmailMultiAlternatives(
                subject,
                text_body,
                settings.DEFAULT_FROM_EMAIL,
                [student.email],
            )
            message.attach_alternative(html_body, 'text/html')
            send_email_async(message, 'إرسال تعليمات الفعالية')
        AdminLog.objects.create(action=f'تم إرسال تعليمات الفعالية لطلاب الفعالية الحالية لعدد {students.count()} طالب')

    threading.Thread(target=_job, daemon=True).start()
    messages.success(request, f'تم بدء عملية إرسال التعليمات لطلاب الفعالية الحالية لـ {students.count()} طالب في الخلفية.')
    return redirect('dashboard:admin_event_settings')


@login_required
def admin_send_certificates(request):
    if not require_admin(request.user):
        return HttpResponseForbidden()
    if request.method != 'POST':
        return redirect('dashboard:admin_reports')
    try:
        from weasyprint import HTML
    except Exception as e:
        AdminLog.objects.create(
            action=f'فشل تحميل مكتبة توليد ملفات PDF (weasyprint): {e}',
        )
        messages.error(
            request,
            'مكتبة توليد ملفات PDF غير متوفرة بالكامل على الخادم. '
            'برجاء مراجعة مسؤول الخادم لتثبيت التبعيات اللازمة.',
        )
        return redirect('dashboard:admin_reports')
    event = Event.get_current()
    if not event.is_finished:
        messages.error(request, 'لا يمكن إصدار شهادات الحضور قبل إنهاء الفعالية.')
        return redirect('dashboard:admin_reports')
    total_checked_in = Student.objects.filter(checked_in=True).count()
    
    # فلترة الطلاب الذين سجلوا حضور الفعالية فقط
    students_eligible = (
        Student.objects.filter(checked_in=True, email__isnull=False, is_certificate_banned=False)
        .exclude(email='')
        .select_related('group')
    )
    
    if not students_eligible.exists():
        if total_checked_in:
            messages.warning(
                request,
                f'لم يتم العثور على طلاب مستحقين للشهادة. تأكد من أن الطلاب قد حضروا جميع جلسات ورش العمل الخاصة بمجموعاتهم.',
            )
        else:
            messages.error(request, 'لا يوجد طلاب تم تسجيل حضورهم لإرسال الشهادات.')
        return redirect('dashboard:admin_reports')
    
    student_ids = list(students_eligible.values_list('id', flat=True))
    count = len(student_ids)
    AdminLog.objects.create(
        action=f'تم بدء عملية إرسال شهادات الحضور PDF في الخلفية لعدد {count} طالبًا (الذين أتموا حضور كافة الورش).',
    )

    def _send_certificates_job(event_id, student_id_list):
        from weasyprint import HTML
        event_local = Event.objects.get(pk=event_id)
        sent_count_local = 0
        skipped_count_local = 0
        for student in Student.objects.filter(id__in=student_id_list).select_related('group'):
            try:
                present_count = 1
                attended_workshops = []
                
                student_identifier = student.student_id or student.id
                qr_payload = f'https://verify.edutech-egy.com/td/{student_identifier}'
                context = {
                    'student': student,
                    'event': event_local,
                    'present_count': present_count,
                    'attended_workshops': attended_workshops,
                    'qr_payload': qr_payload,
                }
                html_string = render_to_string('students/certificate.html', context)
                pdf_bytes = HTML(
                    string=html_string,
                    base_url=settings.SITE_BASE_URL,
                ).write_pdf()
                subject = 'شهادة تقدير – فعالية Tech Day – EduTech Egypt'
                name = student.name
                safe_name = ''.join(ch for ch in name if ch.isalnum() or ch in (' ', '-', '_')).strip() or f'Student-{student.id}'
                filename = f'{safe_name}.pdf'
                text_body = (
                    f'مرحبًا {name},\n\n'
                    f'مرفق مع هذه الرسالة شهادة تقدير لمشاركتك في فعالية Tech Day – الفريق التقني بالقليوبية.\n\n'
                    f'تحياتنا،\n'
                    f'EduTech Egypt System'
                )
                
                content_blocks = """
                    <div class="td-email-box" style="padding:20px;border-radius:20px;background-color:#0f172a;border:1px solid #1e293b;margin-bottom:20px;text-align:center;">
                      <p class="td-email-text-main" style="margin:0 0 10px;font-size:14px;color:#94a3b8;">
                        نشكرك على مشاركتك وحضورك فعاليات Tech Day – الفريق التقني بالقليوبية.
                      </p>
                    </div>
                    <div class="td-email-box" style="padding:20px;border-radius:20px;background-color:#0f172a;border:1px solid #1e293b;text-align:center;">
                      <p class="td-email-text-main" style="margin:0 0 15px;font-size:14px;color:#e5e7eb;font-weight:700;">📜 شهادة التقدير</p>
                      <p class="td-email-text-main" style="margin:0;font-size:13px;color:#cbd5f5;line-height:1.6;">
                        تجد مرفقاً مع هذا البريد شهادة رسمية بصيغة PDF توثق مشاركتك وإنجازك.<br>
                        نتمنى لك كل التوفيق في مسيرتك التقنية القادمة.
                      </p>
                    </div>
                """
                
                html_body = get_styled_email_html(
                    subject=subject,
                    preview_text=f"شهادة تقديرك في فعالية Tech Day - {name}",
                    title="🎓 مبارك النجاح والإنجاز",
                    main_text=f"مرحبًا {name}، يسعدنا تسليمك شهادة مشاركتك في الفعالية.",
                    content_blocks_html=content_blocks
                )
                
                message = EmailMultiAlternatives(
                    subject,
                    text_body,
                    settings.DEFAULT_FROM_EMAIL,
                    [student.email],
                )
                message.attach_alternative(html_body, 'text/html')
                message.attach(filename, pdf_bytes, 'application/pdf')
                send_email_async(message, 'إرسال شهادة حضور PDF')
                sent_count_local += 1
            except Exception as e:
                skipped_count_local += 1
                AdminLog.objects.create(
                    action=f'فشل إنشاء أو إرسال شهادة حضور للطالب {student.id} ({student.email}): {e}',
                )
        AdminLog.objects.create(
            action=f'اكتملت عملية إرسال شهادات الحضور PDF لعدد {sent_count_local} طالبًا (تخطي {skipped_count_local})',
        )

    threading.Thread(
        target=_send_certificates_job,
        args=(event.id, student_ids),
        daemon=True,
    ).start()
    messages.success(
        request,
        f'تم بدء عملية إرسال شهادات الحضور PDF في الخلفية لعدد {count} طالبًا. يمكنك متابعة السجلات من لوحة التحكم.',
    )
    return redirect('dashboard:admin_reports')


@login_required
def admin_public_screen(request):
    if not require_admin(request.user):
        return HttpResponseForbidden()
    screen_url = request.build_absolute_uri('/')
    latest_notification = Notification.objects.filter(is_active=True).first()
    return render(
        request,
        'dashboard/admin_public_screen.html',
        {'screen_url': screen_url, 'latest_notification': latest_notification},
    )


@login_required
def admin_logs(request):
    if not require_admin(request.user):
        return HttpResponseForbidden()
    logs = AdminLog.objects.all()[:500]
    return render(request, 'dashboard/admin_logs.html', {'logs': logs})


@login_required
def admin_custom_email(request):
    if not require_admin(request.user):
        return HttpResponseForbidden()
    initial_to = ''
    groups = Group.objects.all()
    if request.method == 'POST':
        fill_action = (request.POST.get('fill') or '').strip()
        current_to = (request.POST.get('to') or '').strip()
        
        if fill_action:
            # Parse existing emails while preserving order and casing
            existing_list = [e.strip() for e in current_to.replace(';', ',').split(',') if e.strip()]
            existing_lower = {e.lower() for e in existing_list}
            new_emails = []

            if fill_action == 'students':
                emails_qs = Student.objects.filter(email__isnull=False).exclude(email='')
                new_emails = list(emails_qs.values_list('email', flat=True))
                msg_success = f'تم إضافة {len(new_emails)} طالبًا للقائمة.'
            
            elif fill_action == 'supervisors':
                emails_qs = User.objects.filter(role=User.Roles.SUPERVISOR, email__isnull=False).exclude(email='')
                new_emails = list(emails_qs.values_list('email', flat=True))
                msg_success = f'تم إضافة {len(new_emails)} مشرفًا للقائمة.'
            
            elif fill_action == 'group':
                group_id = request.POST.get('preset_group_id') or ''
                if not group_id:
                    messages.error(request, 'برجاء اختيار مجموعة أولاً.')
                else:
                    emails_qs = Student.objects.filter(group_id=group_id, email__isnull=False).exclude(email='')
                    new_emails = list(emails_qs.values_list('email', flat=True))
                    msg_success = f'تم إضافة {len(new_emails)} طالبًا من المجموعة.'
            
            elif fill_action == 'search_code':
                code = (request.POST.get('student_code_search') or '').strip()
                if not code:
                    messages.error(request, 'برجاء إدخال كود الطالب.')
                else:
                    student = Student.objects.filter(student_id=code).first()
                    if student:
                        if student.email:
                            new_emails = [student.email]
                            msg_success = f'تم إضافة بريد الطالب: {student.name}'
                        else:
                            messages.warning(request, f'الطالب {student.name} ليس لديه بريد مسجل.')
                    else:
                        messages.error(request, f'لا يوجد طالب بالكود: {code}')

            # Append only non-duplicates and maintain original order
            added_count = 0
            for email in new_emails:
                clean_email = email.strip()
                if clean_email.lower() not in existing_lower:
                    existing_list.append(clean_email)
                    existing_lower.add(clean_email.lower())
                    added_count += 1
            
            initial_to = ', '.join(existing_list)
            if new_emails and 'msg_success' in locals():
                if added_count == 0 and len(new_emails) > 0:
                    messages.info(request, 'هذه العناوين موجودة بالفعل في القائمة.')
                else:
                    messages.success(request, msg_success)
        
        else:
            to_raw = current_to
            subject = (request.POST.get('subject') or '').strip()
            body = (request.POST.get('body') or '').strip()
            recipients = [
                part.strip()
                for part in to_raw.replace(';', ',').split(',')
                if part.strip()
            ]
            if not recipients or not subject or not body:
                messages.error(request, 'برجاء إدخال عناوين البريد، العنوان، ونص الرسالة.')
                initial_to = to_raw
            else:
                text_body = body
                safe_body_html = body.replace('\n', '<br>')
                
                html_body = get_styled_email_html(
                    subject=subject,
                    preview_text=subject,
                    title="📢 رسالة إدارية",
                    main_text=f"إشعار رسمي من إدارة فعالية Tech Day بالقليوبية.",
                    content_blocks_html=f"""
                        <div class="td-email-box" style="padding:24px;border-radius:24px;background-color:#0f172a;border:1px solid #1e293b;">
                          <p class="td-email-text-main" style="margin:0;font-size:15px;color:#e5e7eb;line-height:1.8;">
                            {safe_body_html}
                          </p>
                        </div>
                    """
                )
                
                message = EmailMultiAlternatives(
                    subject,
                    text_body,
                    settings.DEFAULT_FROM_EMAIL,
                    recipients,
                )
                message.attach_alternative(html_body, 'text/html')
                attachment = request.FILES.get('attachment')
                if attachment:
                    content = attachment.read()
                    content_type = attachment.content_type or 'application/octet-stream'
                    message.attach(attachment.name, content, content_type)
                send_email_async(message, 'إرسال بريد مخصص من لوحة التحكم')
                AdminLog.objects.create(
                    action=f'تم جدولة إرسال بريد مخصص إلى {len(recipients)} مستلم(ين).',
                )
                messages.success(
                    request,
                    f'تم جدولة إرسال البريد إلى {len(recipients)} مستلم(ين).',
                )
                return redirect('dashboard:admin_custom_email')
    context = {
        'DEFAULT_FROM_EMAIL': settings.DEFAULT_FROM_EMAIL,
        'groups': groups,
        'initial_to': initial_to,
        'subject': request.POST.get('subject', ''),
        'body': request.POST.get('body', ''),
    }
    return render(request, 'dashboard/admin_custom_email.html', context)


@login_required
def admin_failed_emails_list(request):
    if not require_admin(request.user):
        return HttpResponseForbidden()
    
    failed_emails = FailedEmail.objects.all().order_by('-created_at')
    return render(request, 'dashboard/admin_failed_emails.html', {'failed_emails': failed_emails})


@login_required
def admin_failed_email_retry(request, pk):
    if not require_admin(request.user):
        return HttpResponseForbidden()
    
    failed_email = get_object_or_404(FailedEmail, pk=pk)
    
    try:
        # إنشاء الرسالة من البيانات المخزنة
        message = EmailMultiAlternatives(
            failed_email.subject,
            failed_email.body_text,
            settings.DEFAULT_FROM_EMAIL,
            failed_email.recipient.split(','),
        )
        if failed_email.body_html:
            message.attach_alternative(failed_email.body_html, 'text/html')
        
        # محاولة الإرسال
        message.send(fail_silently=False)
        
        # في حال النجاح، حذف السجل من قاعدة البيانات
        AdminLog.objects.create(action=f'تم إعادة إرسال البريد بنجاح إلى {failed_email.recipient}')
        failed_email.delete()
        messages.success(request, 'تم إعادة إرسال البريد بنجاح.')
    except Exception as e:
        failed_email.retry_count += 1
        failed_email.error_message = str(e)
        failed_email.save()
        messages.error(request, f'فشل الإرسال مرة أخرى: {e}')
    
    return redirect('dashboard:admin_failed_emails_list')


@login_required
def admin_failed_email_delete(request, pk):
    if not require_admin(request.user):
        return HttpResponseForbidden()
    
    failed_email = get_object_or_404(FailedEmail, pk=pk)
    failed_email.delete()
    messages.success(request, 'تم حذف سجل البريد الفاشل.')
    return redirect('dashboard:admin_failed_emails_list')


@login_required
def admin_student_manual_checkin(request, pk):
    if not require_admin(request.user):
        return HttpResponseForbidden()
    if request.method != 'POST':
        return HttpResponseForbidden()
    
    student = get_object_or_404(Student, pk=pk)
    event = Event.get_current()
    now = timezone.localtime()
    
    # 1. تحديث حالة الحضور العامة للطالب
    student.checked_in = True
    student.checked_in_at = now
    student.save(update_fields=['checked_in', 'checked_in_at'])
    
    # 2. تحديث حالة الحضور في إحصائيات الفعالية الحالية
    stats, _ = StudentEventStats.objects.get_or_create(student=student, event=event)
    stats.checked_in = True
    stats.checked_in_at = now
    stats.save(update_fields=['checked_in', 'checked_in_at'])
    
    # 3. محاولة تسجيل حضور لأي جلسة نشطة حالياً أو أقرب جلسة
    session = WorkshopSession.objects.filter(
        group=student.group,
        start_time__lte=now.time(),
        end_time__gte=now.time(),
    ).first()
    
    if not session and student.group:
        # إذا لم توجد جلسة نشطة، نأخذ أول جلسة لمجموعته في الفعالية لمنحه نقاطها الأساسية
        session = WorkshopSession.objects.filter(group=student.group).order_by('start_time').first()
    
    if session:
        Attendance.objects.update_or_create(
            student=student,
            session=session,
            defaults={'status': Attendance.Status.PRESENT, 'scanned_at': now}
        )
        # منح النقاط
        points = session.workshop.points_per_session
        student.points += points
        student.save(update_fields=['points'])
        stats.points += points
        stats.save(update_fields=['points'])
        if student.group and event.allow_group_points:
            student.group.points += points
            student.group.save(update_fields=['points'])

    AdminLog.objects.create(
        action=f'تم تحضير الطالب {student.name} يدوياً بواسطة الأدمن {request.user.username}',
        event=event
    )
    
    messages.success(request, f'تم تسجيل حضور الطالب {student.name} بنجاح.')
    return redirect(request.META.get('HTTP_REFERER', 'dashboard:admin_current_event_students'))


@login_required
def admin_clearance_all(request):
    if not require_admin(request.user):
        return HttpResponseForbidden()
    
    event = Event.get_current()
    # جلب الطلاب الذين سجلوا حضوراً في الفعالية الحالية فقط
    current_students_qs = Student.objects.filter(
        registrations__event=event,
        registrations__status=StudentRegistration.Status.APPROVED,
        registrations__removed_at__isnull=True,
        checked_in=True
    ).distinct().order_by('name')
    
    # تجهيز الأسماء (أول 3 أسماء فقط)
    students_list = []
    for student in current_students_qs:
        name_parts = (student.name or "").split()
        student.short_name = " ".join(name_parts[:3])
        students_list.append(student)
    
    return render(request, 'dashboard/admin_clearance_all.html', {
        'students': students_list,
        'event': event
    })


@login_required
def admin_devices(request):
    if not require_admin(request.user):
        return HttpResponseForbidden()

    if request.method == 'POST':
        action = (request.POST.get('action') or '').strip()
        device_pk = request.POST.get('device_pk')
        reason = (request.POST.get('reason') or '').strip()
        device = get_object_or_404(MobileDevice, pk=device_pk)
        now = timezone.localtime()
        if action == 'ban':
            device.is_banned = True
            device.banned_reason = reason
            device.banned_at = now
            device.banned_by = request.user
            device.save(update_fields=['is_banned', 'banned_reason', 'banned_at', 'banned_by', 'updated_at'])
            if device.fcm_device_id:
                FCMDevice.objects.filter(id=device.fcm_device_id).update(active=False)
            AdminLog.objects.create(
                action=f'تم حظر جهاز {device.device_id} للمستخدم {device.user.username if device.user else "-"}',
            )
            messages.success(request, 'تم حظر الجهاز.')
        elif action == 'unban':
            device.is_banned = False
            device.banned_reason = ''
            device.banned_at = None
            device.banned_by = None
            device.save(update_fields=['is_banned', 'banned_reason', 'banned_at', 'banned_by', 'updated_at'])
            if device.fcm_device_id:
                FCMDevice.objects.filter(id=device.fcm_device_id).update(active=True)
            AdminLog.objects.create(
                action=f'تم إلغاء حظر جهاز {device.device_id} للمستخدم {device.user.username if device.user else "-"}',
            )
            messages.success(request, 'تم إلغاء حظر الجهاز.')
        elif action == 'disable_notifications':
            if device.fcm_device_id:
                FCMDevice.objects.filter(id=device.fcm_device_id).update(active=False)
                messages.success(request, 'تم تعطيل إشعارات هذا الجهاز.')
        elif action == 'enable_notifications':
            if not device.is_banned and device.fcm_device_id:
                FCMDevice.objects.filter(id=device.fcm_device_id).update(active=True)
                messages.success(request, 'تم تفعيل إشعارات هذا الجهاز.')
        return redirect(request.META.get('HTTP_REFERER', 'dashboard:admin_devices'))

    q = (request.GET.get('q') or '').strip()
    status_filter = (request.GET.get('status') or '').strip().lower() or 'all'
    devices_qs = MobileDevice.objects.select_related('user', 'banned_by').order_by('-last_seen_at')
    if q:
        devices_qs = devices_qs.filter(
            Q(device_id__icontains=q)
            | Q(name__icontains=q)
            | Q(device_model__icontains=q)
            | Q(manufacturer__icontains=q)
            | Q(os_version__icontains=q)
            | Q(user__username__icontains=q)
            | Q(user__email__icontains=q)
        )
    if status_filter == 'banned':
        devices_qs = devices_qs.filter(is_banned=True)
    elif status_filter == 'allowed':
        devices_qs = devices_qs.filter(is_banned=False)

    total_count = MobileDevice.objects.count()
    banned_count = MobileDevice.objects.filter(is_banned=True).count()
    allowed_count = total_count - banned_count

    paginator = Paginator(devices_qs, 50)
    page_obj = paginator.get_page(request.GET.get('page') or 1)
    fcm_ids = [d.fcm_device_id for d in page_obj.object_list if d.fcm_device_id]
    fcm_map = {d.id: d for d in FCMDevice.objects.filter(id__in=fcm_ids)}
    for d in page_obj.object_list:
        fcm = fcm_map.get(d.fcm_device_id) if d.fcm_device_id else None
        d.fcm_active = bool(getattr(fcm, 'active', False)) if fcm else False
        d.fcm_type = getattr(fcm, 'type', '') if fcm else ''
        token = getattr(fcm, 'registration_id', '') if fcm else ''
        d.fcm_token_short = (token[:18] + '…') if token and len(token) > 18 else token

    return render(
        request,
        'dashboard/admin_devices.html',
        {
            'devices': page_obj,
            'q': q,
            'status': status_filter,
            'total_count': total_count,
            'banned_count': banned_count,
            'allowed_count': allowed_count,
        },
    )


@login_required
def admin_device_detail(request, pk):
    if not require_admin(request.user):
        return HttpResponseForbidden()

    device = get_object_or_404(MobileDevice.objects.select_related('user', 'banned_by'), pk=pk)
    fcm_device = None
    if device.fcm_device_id:
        fcm_device = FCMDevice.objects.filter(id=device.fcm_device_id).first()

    if request.method == 'POST':
        action = (request.POST.get('action') or '').strip()
        reason = (request.POST.get('reason') or '').strip()
        now = timezone.localtime()
        if action == 'ban':
            device.is_banned = True
            device.banned_reason = reason
            device.banned_at = now
            device.banned_by = request.user
            device.save(update_fields=['is_banned', 'banned_reason', 'banned_at', 'banned_by', 'updated_at'])
            if device.fcm_device_id:
                FCMDevice.objects.filter(id=device.fcm_device_id).update(active=False)
            messages.success(request, 'تم حظر الجهاز.')
        elif action == 'unban':
            device.is_banned = False
            device.banned_reason = ''
            device.banned_at = None
            device.banned_by = None
            device.save(update_fields=['is_banned', 'banned_reason', 'banned_at', 'banned_by', 'updated_at'])
            if device.fcm_device_id:
                FCMDevice.objects.filter(id=device.fcm_device_id).update(active=True)
            messages.success(request, 'تم إلغاء حظر الجهاز.')
        elif action == 'disable_notifications':
            if device.fcm_device_id:
                FCMDevice.objects.filter(id=device.fcm_device_id).update(active=False)
                messages.success(request, 'تم تعطيل إشعارات هذا الجهاز.')
        elif action == 'enable_notifications':
            if not device.is_banned and device.fcm_device_id:
                FCMDevice.objects.filter(id=device.fcm_device_id).update(active=True)
                messages.success(request, 'تم تفعيل إشعارات هذا الجهاز.')
        return redirect('dashboard:admin_device_detail', pk=device.pk)

    return render(
        request,
        'dashboard/admin_device_detail.html',
        {'device': device, 'fcm_device': fcm_device},
    )


@login_required
def admin_public_forms(request):
    if not require_admin(request.user):
        return HttpResponseForbidden()

    q = (request.GET.get('q') or '').strip()
    forms_qs = PublicForm.objects.select_related('created_by').annotate(submissions_count=Count('submissions')).order_by('-created_at')
    if q:
        forms_qs = forms_qs.filter(Q(title__icontains=q) | Q(description__icontains=q))

    paginator = Paginator(forms_qs, 50)
    page_obj = paginator.get_page(request.GET.get('page') or 1)
    return render(
        request,
        'dashboard/admin_public_forms.html',
        {'forms': page_obj, 'q': q},
    )


@login_required
def admin_public_form_detail(request, pk):
    if not require_admin(request.user):
        return HttpResponseForbidden()

    form_obj = get_object_or_404(PublicForm.objects.select_related('created_by'), pk=pk)
    if request.method == 'POST':
        action = (request.POST.get('action') or '').strip()
        if action == 'toggle_active':
            form_obj.is_active = not form_obj.is_active
            form_obj.save(update_fields=['is_active'])
            messages.success(request, 'تم تحديث حالة الاستمارة.')
        elif action == 'claim':
            if not form_obj.created_by_id:
                form_obj.created_by = request.user
                form_obj.save(update_fields=['created_by'])
                messages.success(request, 'تم تعيينك كمنشئ للاستمارة.')
        elif action == 'set_custom_link':
            raw_slug = (request.POST.get('custom_slug') or '').strip().lower()
            if not raw_slug:
                form_obj.custom_slug = None
                form_obj.save(update_fields=['custom_slug'])
                messages.success(request, 'تم حذف الرابط المخصص والرجوع للرابط الافتراضي.')
            elif PublicForm.objects.exclude(pk=form_obj.pk).filter(custom_slug=raw_slug).exists():
                messages.error(request, 'هذا الرابط مستخدم بالفعل، اختر رابطًا مختلفًا.')
            else:
                form_obj.custom_slug = raw_slug
                try:
                    form_obj.full_clean()
                except ValidationError:
                    messages.error(request, 'صيغة الرابط غير صالحة. استخدم حروف إنجليزية وأرقام وشرطات فقط.')
                else:
                    form_obj.save(update_fields=['custom_slug'])
                    messages.success(request, 'تم حفظ الرابط المخصص بنجاح.')
        return redirect('dashboard:admin_public_form_detail', pk=form_obj.pk)

    fields = list(form_obj.fields.all())
    submissions_count = form_obj.submissions.count()
    return render(
        request,
        'dashboard/admin_public_form_detail.html',
        {'form_obj': form_obj, 'fields': fields, 'submissions_count': submissions_count},
    )


@login_required
def admin_public_form_secret_link(request, pk):
    if not require_admin(request.user):
        return HttpResponseForbidden()
    form_obj = get_object_or_404(PublicForm, pk=pk)
    public_key = form_obj.custom_slug or form_obj.token
    path = reverse('public_screen:public_form', args=[public_key])
    return JsonResponse({'url': request.build_absolute_uri(path)})


@login_required
def admin_public_form_submissions(request, pk):
    if not require_admin(request.user):
        return HttpResponseForbidden()

    form_obj = get_object_or_404(PublicForm, pk=pk)
    submissions_qs = PublicFormSubmission.objects.filter(form=form_obj).prefetch_related(
        'answers',
        'answers__field',
    )
    paginator = Paginator(submissions_qs, 50)
    page_obj = paginator.get_page(request.GET.get('page') or 1)

    field_keys = {
        'full_name': 'الاسم',
        'student_code': 'كود الطالب',
        'phone': 'الهاتف',
        'education_admin': 'الإدارة',
    }
    for sub in page_obj.object_list:
        by_key = {}
        for ans in sub.answers.all():
            by_key[ans.field.key] = ans.value_text
        sub.summary = {k: by_key.get(k, '') for k in field_keys.keys()}

    return render(
        request,
        'dashboard/admin_public_form_submissions.html',
        {'form_obj': form_obj, 'submissions': page_obj},
    )


@login_required
def admin_public_form_export_excel(request, pk):
    if not require_admin(request.user):
        return HttpResponseForbidden()

    form_obj = get_object_or_404(PublicForm, pk=pk)
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except Exception:
        messages.error(request, 'مكتبة openpyxl غير متوفرة. يرجى تثبيتها: pip install openpyxl')
        return redirect('dashboard:admin_public_form_submissions', pk=form_obj.pk)

    fields = list(form_obj.fields.all().order_by('order', 'label'))
    submissions = list(
        PublicFormSubmission.objects.filter(form=form_obj)
        .prefetch_related('answers', 'answers__field')
        .order_by('-submitted_at')
    )

    wb = Workbook()
    ws = wb.active
    ws.title = 'الردود'
    ws.sheet_view.rightToLeft = True

    header = ['وقت الإرسال', 'IP'] + [f.label for f in fields]
    ws.append(header)

    head_fill = PatternFill(start_color='0F172A', end_color='0F172A', fill_type='solid')
    head_font = Font(color='FFFFFF', bold=True)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    border = Border(
        left=Side(style='thin', color='334155'),
        right=Side(style='thin', color='334155'),
        top=Side(style='thin', color='334155'),
        bottom=Side(style='thin', color='334155'),
    )

    for i in range(1, len(header) + 1):
        cell = ws.cell(row=1, column=i)
        cell.fill = head_fill
        cell.font = head_font
        cell.alignment = center
        cell.border = border

    for submission in submissions:
        answers_map = {}
        for ans in submission.answers.all():
            answers_map.setdefault(ans.field_id, []).append(ans)
        row = [
            timezone.localtime(submission.submitted_at).strftime('%Y-%m-%d %H:%M'),
            submission.ip_address or '',
        ]
        for f in fields:
            ans_list = answers_map.get(f.id) or []
            if not ans_list:
                row.append('')
                continue
            if any(ans.value_file for ans in ans_list):
                file_values = []
                for ans in ans_list:
                    if not ans.value_file:
                        continue
                    try:
                        file_values.append(request.build_absolute_uri(ans.value_file.url))
                    except Exception:
                        file_values.append(ans.value_file.name or '')
                row.append('\n'.join(file_values))
            else:
                text_values = [ans.value_text or '' for ans in ans_list if ans.value_text]
                row.append('\n'.join(text_values))
        ws.append(row)

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(header)):
        for cell in row:
            cell.alignment = Alignment(vertical='top', wrap_text=True)
            cell.border = border

    ws.freeze_panes = 'A2'
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 16
    for idx in range(3, len(header) + 1):
        ws.column_dimensions[get_column_letter(idx)].width = 28

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(
        buffer.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    safe_name = ''.join(ch if ch.isalnum() else '_' for ch in form_obj.title)[:40] or 'public_form'
    response['Content-Disposition'] = f'attachment; filename="{safe_name}_responses.xlsx"'
    return response


@login_required
def admin_public_form_submission_detail(request, form_pk, submission_pk):
    if not require_admin(request.user):
        return HttpResponseForbidden()

    form_obj = get_object_or_404(PublicForm, pk=form_pk)
    submission = get_object_or_404(
        PublicFormSubmission.objects.filter(form=form_obj).prefetch_related('answers', 'answers__field'),
        pk=submission_pk,
    )
    answers = list(submission.answers.all())
    answers.sort(key=lambda a: a.field.order)
    for ans in answers:
        ans.file_exists = False
        ans.file_url = ''
        if ans.value_file:
            file_name = ans.value_file.name or ''
            if file_name:
                ans.file_exists = default_storage.exists(file_name)
                if ans.file_exists:
                    try:
                        ans.file_url = ans.value_file.url
                    except Exception:
                        ans.file_url = ''
    return render(
        request,
        'dashboard/admin_public_form_submission_detail.html',
        {'form_obj': form_obj, 'submission': submission, 'answers': answers},
    )


@login_required
def admin_public_form_submission_delete(request, form_pk, submission_pk):
    if not require_admin(request.user):
        return HttpResponseForbidden()
    if request.method != 'POST':
        return redirect('dashboard:admin_public_form_submissions', pk=form_pk)

    form_obj = get_object_or_404(PublicForm, pk=form_pk)
    submission = get_object_or_404(PublicFormSubmission, pk=submission_pk, form=form_obj)

    for ans in submission.answers.all():
        if ans.value_file:
            ans.value_file.delete(save=False)
    submission.delete()

    messages.success(request, 'تم حذف الرد بنجاح.')
    return redirect('dashboard:admin_public_form_submissions', pk=form_obj.pk)


def media_proxy(request, file_path):
    normalized = (unquote(file_path or '')).lstrip('/')
    if not normalized:
        raise Http404()
    parts = normalized.replace('\\', '/').split('/')
    if any(p in {'..', ''} for p in parts):
        raise Http404()
    if not default_storage.exists(normalized):
        raise Http404()
    stream = default_storage.open(normalized, mode='rb')
    content_type, _ = mimetypes.guess_type(normalized)
    response = FileResponse(stream, content_type=content_type or 'application/octet-stream')
    return response


@login_required
def admin_smtp_diagnose(request):
    if not require_admin(request.user):
        return HttpResponseForbidden()

    host = settings.EMAIL_HOST
    port = int(getattr(settings, 'EMAIL_PORT', 0) or 0)
    use_ssl = bool(getattr(settings, 'EMAIL_USE_SSL', False))
    use_tls = bool(getattr(settings, 'EMAIL_USE_TLS', False))
    timeout = int(getattr(settings, 'EMAIL_TIMEOUT', 60) or 60)
    user = getattr(settings, 'EMAIL_HOST_USER', '') or ''
    has_password = bool(getattr(settings, 'EMAIL_HOST_PASSWORD', '') or '')

    started = time.monotonic()
    try:
        # Create a more robust SSL context
        try:
            import certifi
            context = ssl.create_default_context(cafile=certifi.where())
        except (ImportError, Exception):
            context = ssl.create_default_context()
        
        # Check for bypass
        if os.environ.get('SMTP_INSECURE_SSL', '').strip() == '1':
            context = ssl._create_unverified_context()

        if use_ssl:
            server = smtplib.SMTP_SSL(host=host, port=port, timeout=timeout, context=context)
        else:
            server = smtplib.SMTP(host=host, port=port, timeout=timeout)
        server.ehlo()
        if use_tls:
            server.starttls(context=context)
            server.ehlo()
        if user and has_password:
            server.login(user, settings.EMAIL_HOST_PASSWORD)
        server.quit()
        elapsed_ms = int((time.monotonic() - started) * 1000)
        return JsonResponse(
            {
                'ok': True,
                'host': host,
                'port': port,
                'use_ssl': use_ssl,
                'use_tls': use_tls,
                'timeout_seconds': timeout,
                'login_attempted': bool(user and has_password),
                'elapsed_ms': elapsed_ms,
            }
        )
    except Exception as e:
        elapsed_ms = int((time.monotonic() - started) * 1000)
        return JsonResponse(
            {
                'ok': False,
                'host': host,
                'port': port,
                'use_ssl': use_ssl,
                'use_tls': use_tls,
                'timeout_seconds': timeout,
                'login_attempted': bool(user and has_password),
                'elapsed_ms': elapsed_ms,
                'error': str(e),
            },
            status=502,
        )
